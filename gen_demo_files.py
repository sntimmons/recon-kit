import csv, random
from pathlib import Path
from datetime import date, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XL = True
except ImportError:
    HAS_XL = False

random.seed(42)
OUT = Path('c:/Users/steph/recon-kit/site/downloads')
OUT.mkdir(exist_ok=True)

FIRST = ['James','Mary','Robert','Patricia','John','Jennifer','Michael','Linda',
         'William','Barbara','David','Susan','Richard','Jessica','Joseph','Sarah',
         'Thomas','Karen','Charles','Lisa','Christopher','Nancy','Daniel','Betty',
         'Matthew','Sandra','Anthony','Emily','Mark','Donna','Donald','Sharon',
         'Steven','Kimberly','Paul','Melissa','Andrew','Carol','Joshua','Amanda',
         'Kenneth','Deborah','Kevin','Stephanie','Brian','Rebecca','George','Laura',
         'Timothy','Michelle','Ronald','Cynthia','Edward','Ashley','Jason','Dorothy']
LAST  = ['Smith','Johnson','Williams','Brown','Jones','Garcia','Miller','Davis',
         'Rodriguez','Martinez','Hernandez','Lopez','Gonzalez','Wilson','Anderson',
         'Thomas','Taylor','Moore','Jackson','Martin','Lee','Perez','Thompson',
         'White','Harris','Sanchez','Clark','Ramirez','Lewis','Robinson','Walker',
         'Young','Allen','King','Wright','Scott','Torres','Nguyen','Hill','Flores',
         'Green','Adams','Nelson','Baker','Hall','Rivera','Campbell','Mitchell',
         'Carter','Roberts','Okafor','Marchetti','Fontaine','Nakamura','Vasquez']
POSITIONS = ['HR Manager','Software Engineer','Data Analyst','Project Manager',
             'Financial Analyst','Operations Coordinator','Marketing Specialist',
             'IT Support Specialist','Accountant','HR Business Partner',
             'Product Manager','Business Analyst','Recruiter','Payroll Specialist',
             'Systems Administrator','Compliance Officer','Training Coordinator',
             'Benefits Administrator','UX Designer','Sales Representative']
DISTRICTS = ['Northeast','Southeast','Midwest','Southwest','Northwest','Mid-Atlantic',
              'Great Lakes','Mountain West','Pacific Coast','Central']
STATES = ['CA','TX','NY','FL','IL','PA','OH','GA','NC','MI','NJ','VA','WA','AZ','MA']
STATUSES = ['Active']*8 + ['Inactive','Leave']
SALARIES = [45000,52000,58000,62000,68000,74000,80000,88000,95000,105000,115000,125000]

def rand_hire():
    return date(random.randint(2010,2023), random.randint(1,12), random.randint(1,28))

rows = []
for i in range(244):
    fn = random.choice(FIRST); ln = random.choice(LAST)
    name = f'{fn} {ln}'
    wid  = f'EMP-{10100+i:05d}'
    sal  = random.choice(SALARIES)
    hire = rand_hire()
    if i < 188:
        ms = 'worker_id'
    elif i < 224:
        ms = 'last4_dob'
    elif i < 238:
        ms = 'dob_name'
    else:
        ms = 'name_hire_date'
    conf = '' if ms == 'worker_id' else round(random.uniform(0.89, 0.99), 4)
    pos  = random.choice(POSITIONS)
    dist = random.choice(DISTRICTS)
    st   = random.choice(STATES)
    stat = random.choice(STATUSES)
    rows.append({
        'pair_id': f'P-{i+1:04d}',
        'match_source': ms,
        'confidence': conf,
        'old_worker_id': wid, 'new_worker_id': wid,
        'old_full_name_norm': name.lower(), 'new_full_name_norm': name.lower(),
        'old_worker_status': stat, 'new_worker_status': stat,
        'old_hire_date': hire.isoformat(), 'new_hire_date': hire.isoformat(),
        'old_position': pos, 'new_position': pos,
        'old_district': dist, 'new_district': dist,
        'old_location_state': st, 'new_location_state': st,
        'old_salary': sal, 'new_salary': sal,
        'salary_delta': 0, 'salary_ratio': 1.0,
        'status_changed': False, 'hire_date_changed': False,
        'has_salary_mismatch': False,
        'action': 'APPROVE', 'reason': 'auto_approve',
        'fix_types': '', 'summary': 'clean', 'priority_score': 0,
    })

for idx in [4,11,27,44,66,88,101,133,155,177,200,219,234]:
    old = rows[idx]['old_salary']
    pct = random.choice([1.03,1.05,1.08,1.12,0.97,0.95])
    new = int(old * pct)
    delta = new - old
    rows[idx].update({'new_salary': new, 'salary_delta': delta,
                      'salary_ratio': round(new/old,6), 'has_salary_mismatch': True,
                      'fix_types': 'salary', 'summary': f'salary {old:,} -> {new:,}',
                      'priority_score': 15 if abs(delta)<1000 else 35 if abs(delta)<5000 else 65})

for idx in [7,32,71,97,144,166,198,217]:
    old_hd = date.fromisoformat(rows[idx]['old_hire_date'])
    new_hd = old_hd + timedelta(days=random.choice([1,7,14,30,-7]))
    rows[idx]['new_hire_date'] = new_hd.isoformat()
    rows[idx]['hire_date_changed'] = True
    ft = rows[idx]['fix_types']
    rows[idx]['fix_types'] = (ft+'|' if ft else '') + 'hire_date'
    rows[idx]['priority_score'] = max(rows[idx]['priority_score'], 20)

for idx in [18,54]:
    rows[idx]['new_worker_status'] = 'Inactive' if rows[idx]['old_worker_status']=='Active' else 'Active'
    rows[idx]['status_changed'] = True
    ft = rows[idx]['fix_types']
    rows[idx]['fix_types'] = (ft+'|' if ft else '') + 'status'
    rows[idx]['priority_score'] = max(rows[idx]['priority_score'], 50)

review_rows = [r for r in rows if r['match_source'] != 'worker_id'][:6]
for r in review_rows:
    if not r['has_salary_mismatch']:
        old = r['old_salary']
        new = int(old * random.choice([1.04, 1.07, 0.95]))
        r['new_salary'] = new; r['salary_delta'] = new - old
        r['salary_ratio'] = round(new/old, 6); r['has_salary_mismatch'] = True
        r['fix_types'] = 'salary'; r['summary'] = f'salary {old:,} -> {new:,}'
    r['action'] = 'REVIEW'
    r['reason'] = 'below_confidence_threshold'
    r['confidence'] = round(random.uniform(0.72, 0.84), 4)
    r['priority_score'] = max(r['priority_score'], 50) + random.randint(0, 30)

approved_sal = [r for r in rows if r['has_salary_mismatch'] and r['action']=='APPROVE'][:3]

# wide_compare.csv
WC_FIELDS = ['pair_id','match_source','confidence','action','reason','fix_types','summary',
             'priority_score','old_worker_id','new_worker_id','old_full_name_norm',
             'new_full_name_norm','old_worker_status','new_worker_status',
             'old_hire_date','new_hire_date','old_position','new_position',
             'old_district','new_district','old_location_state','new_location_state',
             'old_salary','new_salary','salary_delta','salary_ratio',
             'status_changed','hire_date_changed','has_salary_mismatch']
with open(OUT/'wide_compare.csv','w',newline='',encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=WC_FIELDS, extrasaction='ignore')
    w.writeheader(); w.writerows(rows)

# review_queue.csv
RQ_FIELDS = ['pair_id','match_source','confidence','old_worker_id','new_worker_id',
             'old_full_name_norm','new_full_name_norm','old_salary','new_salary',
             'salary_delta','old_hire_date','new_hire_date','old_worker_status',
             'new_worker_status','action','reason','priority_score','fix_types','summary']
with open(OUT/'review_queue.csv','w',newline='',encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=RQ_FIELDS, extrasaction='ignore')
    w.writeheader()
    for r in review_rows:
        w.writerow(r)

# corrections_salary.csv
CS_FIELDS = ['worker_id','effective_date','compensation_amount','currency','reason',
             'pair_id','match_source','confidence','summary']
with open(OUT/'corrections_salary.csv','w',newline='',encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=CS_FIELDS)
    w.writeheader()
    for r in approved_sal:
        w.writerow({'worker_id': r['new_worker_id'], 'effective_date': '2026-04-01',
                    'compensation_amount': r['new_salary'], 'currency': 'USD',
                    'reason': 'HRIS migration salary reconciliation - auto approved',
                    'pair_id': r['pair_id'], 'match_source': r['match_source'],
                    'confidence': r['confidence'] if r['confidence'] else 1.0,
                    'summary': r['summary']})

# recon_summary.xlsx
if HAS_XL:
    wb = openpyxl.Workbook()
    HDR_FILL = PatternFill('solid', fgColor='1a1f2e')
    HDR_FONT = Font(bold=True, color='4f8ef7', size=10)
    DEF_FONT = Font(color='c8cdd8', size=10)
    ALT_FILL = PatternFill('solid', fgColor='141820')
    DEF_FILL = PatternFill('solid', fgColor='1a1f2e')

    def style_ws(ws, headers, widths=None):
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal='left')
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = ALT_FILL if ri % 2 == 0 else DEF_FILL
            for cell in row:
                cell.font = DEF_FONT; cell.fill = fill
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        if widths:
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

    ws1 = wb.active; ws1.title = 'Summary'
    ws1.sheet_view.showGridLines = False
    for row in [['Metric','Value'],['Run date','2026-03-05'],['Legacy records',250],
                ['New system records',247],['Total matched',244],['Auto-approved',238],
                ['Sent to review queue',6],['Match rate','97.6%'],
                ['Salary mismatches',13],['Hire date mismatches',8],['Status mismatches',2],
                ['Sanity gate','PASS'],['Engine version','v1.0']]:
        ws1.append(row)
    style_ws(ws1, ['Metric','Value'], {'A':28,'B':20})

    ws2 = wb.create_sheet('All_Matches')
    ws2.sheet_view.showGridLines = False
    h2 = ['pair_id','match_source','confidence','old_worker_id','new_worker_id',
          'old_full_name_norm','old_salary','new_salary','salary_delta',
          'old_hire_date','new_hire_date','old_worker_status','new_worker_status',
          'action','fix_types','priority_score']
    ws2.append(h2)
    for r in rows:
        ws2.append([r['pair_id'],r['match_source'],r['confidence'],r['old_worker_id'],
                    r['new_worker_id'],r['old_full_name_norm'],r['old_salary'],r['new_salary'],
                    r['salary_delta'],r['old_hire_date'],r['new_hire_date'],
                    r['old_worker_status'],r['new_worker_status'],r['action'],
                    r['fix_types'],r['priority_score']])
    style_ws(ws2, h2, {'A':10,'B':16,'C':12,'D':14,'E':14,'F':26,'G':12,'H':12,'I':12,'J':14,'K':14,'L':14,'M':16,'N':10,'O':20,'P':14})

    ws3 = wb.create_sheet('Salary_Mismatches')
    ws3.sheet_view.showGridLines = False
    h3 = ['pair_id','old_worker_id','old_full_name_norm','old_salary','new_salary','salary_delta','salary_ratio','action','confidence']
    ws3.append(h3)
    for r in rows:
        if r['has_salary_mismatch']:
            ws3.append([r['pair_id'],r['old_worker_id'],r['old_full_name_norm'],
                        r['old_salary'],r['new_salary'],r['salary_delta'],r['salary_ratio'],r['action'],r['confidence']])
    style_ws(ws3, h3, {'A':10,'B':14,'C':26,'D':12,'E':12,'F':12,'G':12,'H':10,'I':12})

    ws4 = wb.create_sheet('Status_Mismatches')
    ws4.sheet_view.showGridLines = False
    h4 = ['pair_id','old_worker_id','old_full_name_norm','old_worker_status','new_worker_status','action']
    ws4.append(h4)
    for r in rows:
        if r['status_changed']:
            ws4.append([r['pair_id'],r['old_worker_id'],r['old_full_name_norm'],r['old_worker_status'],r['new_worker_status'],r['action']])
    style_ws(ws4, h4, {'A':10,'B':14,'C':26,'D':16,'E':16,'F':10})

    ws5 = wb.create_sheet('HireDate_Mismatches')
    ws5.sheet_view.showGridLines = False
    h5 = ['pair_id','old_worker_id','old_full_name_norm','old_hire_date','new_hire_date','action']
    ws5.append(h5)
    for r in rows:
        if r['hire_date_changed']:
            ws5.append([r['pair_id'],r['old_worker_id'],r['old_full_name_norm'],r['old_hire_date'],r['new_hire_date'],r['action']])
    style_ws(ws5, h5, {'A':10,'B':14,'C':26,'D':14,'E':14,'F':10})

    ws6 = wb.create_sheet('JobOrg_Mismatches')
    ws6.sheet_view.showGridLines = False
    h6 = ['pair_id','old_worker_id','old_full_name_norm','old_position','new_position','old_district','new_district','action']
    ws6.append(h6)
    style_ws(ws6, h6, {'A':10,'B':14,'C':26,'D':22,'E':22,'F':16,'G':16,'H':10})

    ws7 = wb.create_sheet('Review_Queue')
    ws7.sheet_view.showGridLines = False
    h7 = ['pair_id','old_worker_id','old_full_name_norm','match_source','confidence',
          'old_salary','new_salary','salary_delta','reason','priority_score','fix_types']
    ws7.append(h7)
    for r in review_rows:
        ws7.append([r['pair_id'],r['old_worker_id'],r['old_full_name_norm'],r['match_source'],
                    r['confidence'],r['old_salary'],r['new_salary'],r['salary_delta'],
                    r['reason'],r['priority_score'],r['fix_types']])
    style_ws(ws7, h7, {'A':10,'B':14,'C':26,'D':14,'E':12,'F':12,'G':12,'H':12,'I':32,'J':14,'K':16})

    ws8 = wb.create_sheet('Corrections_Manifest')
    ws8.sheet_view.showGridLines = False
    h8 = ['pair_id','worker_id','fix_type','action','confidence','effective_date','old_value','new_value','reason','match_source']
    ws8.append(h8)
    for r in approved_sal:
        ws8.append([r['pair_id'],r['new_worker_id'],'salary','APPROVE',
                    r['confidence'] if r['confidence'] else 1.0,'2026-04-01',
                    r['old_salary'],r['new_salary'],
                    'HRIS migration salary reconciliation - auto approved',r['match_source']])
    style_ws(ws8, h8, {'A':10,'B':14,'C':12,'D':10,'E':12,'F':14,'G':12,'H':12,'I':48,'J':16})

    wb.save(OUT/'recon_summary.xlsx')
    print('xlsx: OK')
else:
    print('xlsx: SKIPPED - openpyxl not found')

print('wide_compare.csv:', sum(1 for _ in open(OUT/'wide_compare.csv'))-1, 'rows')
print('review_queue.csv:', sum(1 for _ in open(OUT/'review_queue.csv'))-1, 'rows')
print('corrections_salary.csv:', sum(1 for _ in open(OUT/'corrections_salary.csv'))-1, 'rows')

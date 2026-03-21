"""
build_internal_audit_workbook.py - Excel workbook export for internal audit runs.

Sheets (in order):
  1. Executive_Summary  - all findings, mirrors PDF; first tab HR managers see
  2. Fix_List           - action table: every issue, priority-ranked, Blocking? column
  3. Fix_List_Detail    - row-level execution sheet across all actionable issues
  4. Findings_*         - one tab per issue type that has row-level data
  5. Findings_Index     - master index with detail-sheet cross-references
  6. Data_Quality_Score - total records, completeness %, gate status, severity counts

Design rules:
  • Counts + severity come exclusively from internal_audit_report.json (same source as PDF).
  • CRITICAL finding detail: full row expansion from source data.
  • HIGH/MEDIUM finding detail: sample rows from internal_audit_data.csv + note.
  • Dataset-level findings (status_no_terminated): no detail tab.
  • No technical check_key labels visible to HR users.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("[error] openpyxl not installed", file=sys.stderr)
    sys.exit(2)

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from audit import internal_audit as ia


# ── Row caps ──────────────────────────────────────────────────────────────────
DETAIL_ROW_CAP = 50_000   # max rows per Findings_* detail sheet

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL = "D9EAF7"

SEV_FILLS: dict[str, str] = {
    "CRITICAL": "FFCCCC",   # red
    "HIGH":     "FFE0CC",   # orange
    "MEDIUM":   "FFF3CC",   # yellow
    "LOW":      "E8F5E9",   # light green
}

# ── Priority system ───────────────────────────────────────────────────────────
SEVERITY_PRIORITY: dict[str, int] = {
    "CRITICAL": 1,
    "HIGH":     2,
    "MEDIUM":   3,
    "LOW":      4,
}
BLOCKING_SEVERITIES = {"CRITICAL", "HIGH"}

# ── Dataset-level check_keys ──────────────────────────────────────────────────
# For these, count=1 means "the dataset has this issue", not "1 employee affected."
DATASET_LEVEL_CHECKS: set[str] = {"status_no_terminated"}

# ── Check metadata lookup ─────────────────────────────────────────────────────
# Maps check_key → (issue_name, business_impact, required_action, detail_sheet | None)
# detail_sheet = None → dataset-level, no per-row tab created.
_M = {
    # CRITICAL
    "duplicate_worker_id": (
        "Duplicate Worker ID",
        "Data for the wrong employee could be modified or payroll routed incorrectly during migration.",
        "Identify each duplicated ID, assign a unique Worker ID to each employee, and update all references before use.",
        "Findings_Dup_WorkerID",
    ),
    "duplicate_canonical_worker_id_conflict": (
        "Duplicate Worker ID (Source Column Conflict)",
        "Two source columns map to Worker ID with conflicting values - only one can be correct.",
        "Decide which source column is authoritative, remove the other, and ensure each employee has exactly one Worker ID.",
        "Findings_Dup_WorkerID",   # merged into the same detail tab
    ),
    "active_zero_salary": (
        "Missing or Invalid Salary",
        "Active employees will not receive pay, causing immediate payroll failures.",
        "Enter a valid positive salary or payrate for every flagged active employee before any payroll run.",
        "Findings_Missing_Salary",
    ),
    "pay_type_missing_or_invalid": (
        "Missing or Invalid Pay Type",
        "Payroll cannot reliably determine whether to use salary or pay rate logic for the worker.",
        "Populate a valid pay type from the controlled allowed list before payroll or migration.",
        "Findings_Pay_Type",
    ),
    "compensation_type_mismatch": (
        "Compensation Type Mismatch",
        "The worker is missing the required compensation field for the stated pay type, creating payroll setup risk.",
        "Align pay type with the required salary or pay rate field before payroll or migration.",
        "Findings_Comp_Type",
    ),
    "hourly_implausible_payrate": (
        "Implausible Hourly Pay Rate",
        "Hourly compensation is not valid for payroll processing or falls outside the configured review range.",
        "Correct zero or negative pay rates immediately and review outlier rates before payroll or migration.",
        "Findings_Hourly_Rate",
    ),
    "salaried_implausible_salary": (
        "Implausible Annual Salary",
        "Annual salary is not valid for payroll processing or falls outside the configured review range.",
        "Correct zero or negative salaries immediately and review outlier salaries before payroll or migration.",
        "Findings_Impl_Salary",
    ),
    "invalid_date_logic": (
        "Invalid Dates",
        "Service calculations, payroll timing, and benefits eligibility will produce wrong results.",
        "Correct the hire and termination dates for every flagged record so the employment timeline is valid.",
        "Findings_Invalid_Dates",
    ),
    "active_with_termination_date": (
        "Active Employees with Termination Dates",
        "Employees may be excluded from payroll or benefits workflows despite being actively employed.",
        "For each flagged employee, either remove the termination date or change their status to Terminated.",
        "Findings_Active_Term",
    ),
    "missing_required_identity": (
        "Missing Required Identity Fields",
        "Employees cannot be reliably identified or matched in the target HRIS system.",
        "Populate Worker ID, First Name, and Last Name for every employee before any system load.",
        "Findings_Missing_ID",
    ),
    # HIGH
    "phone_invalid": (
        "Invalid Phone Numbers",
        "Employees cannot be contacted for benefits, emergencies, or onboarding - and corrupted phone data suggests other fields may also be affected.",
        "Re-export the file from the source system and verify the phone field mapping is correct.",
        "Findings_Invalid_Phone",
    ),
    "salary_suspicious_default": (
        "Suspicious Salary Defaults",
        "Payroll will use incorrect compensation amounts, causing underpayment or overpayment.",
        "Verify the correct salary for each flagged employee from HR records or payroll before migrating.",
        "Findings_Salary_Defaults",
    ),
    "comp_dual_value_conflict": (
        "Salary and Pay Rate Conflict",
        "The worker record has conflicting compensation values and payroll teams may apply the wrong one.",
        "Review the worker record and keep only the compensation field that matches the intended pay type.",
        "Findings_Comp_Conflict",
    ),
    "missing_standard_hours_hourly": (
        "Missing Standard Hours for Hourly Worker",
        "Payroll teams cannot confidently annualize or validate hourly compensation without standard hours.",
        "Populate standard hours for each hourly worker before migration.",
        "Findings_Std_Hours",
    ),
    "hire_date_suspicious_default": (
        "Suspicious Hire Dates",
        "Seniority, benefits eligibility, and tenure calculations will be wrong for every flagged employee.",
        "Research the correct hire date for each flagged record from offer letters or original employment records.",
        "Findings_Hire_Defaults",
    ),
    "impossible_dates": (
        "Impossible Dates",
        "Tenure, benefits, and compliance calculations will fail or produce inaccurate results.",
        "Correct each impossible date by reviewing the employee's source records.",
        "Findings_Impossible_Dates",
    ),
    "manager_loop": (
        "Manager Reporting Loops",
        "The organisational hierarchy is circular - automated approval workflows will deadlock.",
        "Review each loop and correct the manager assignment so every path ends at a top-level employee with no manager.",
        "Findings_Mgr_Loops",
    ),
    "pay_equity_flag": (
        "Pay Equity Variance",
        "Salary differences above 30% within the same role may indicate inequitable pay that creates legal exposure.",
        "Review each flagged role-department group and document a legitimate reason for every variance above 30%.",
        "Findings_Pay_Equity",
    ),
    "ghost_employee_indicator": (
        "Ghost Employee Indicators",
        "Records show characteristics consistent with payroll fraud - active status with no salary, department, or manager.",
        "Investigate each flagged record immediately with direct managers and payroll records.",
        "Findings_Ghost_Employees",
    ),
    "duplicate_canonical_conflict": (
        "Duplicate Field Mapping Conflict",
        "Multiple source columns resolve to the same HRIS field with conflicting values.",
        "Identify the authoritative source column for each field and remove or consolidate the others.",
        None,   # dataset-wide issue, no per-row tab
    ),
    # MEDIUM
    "duplicate_email": (
        "Duplicate Emails",
        "HRIS login, notifications, and self-service access will break for every employee sharing an address.",
        "Assign a unique, valid email address to each employee before loading data into any system.",
        "Findings_Dup_Email",
    ),
    "duplicate_name_different_id": (
        "Duplicate Names with Different IDs",
        "The matching engine may link corrections to the wrong person during reconciliation.",
        "Review each name pair - merge records if they are the same person, or add a distinguishing identifier if they are different people.",
        "Findings_Dup_Names",
    ),
    "status_no_terminated": (
        "No Terminated Employees",
        "Employment history appears incomplete, which affects compliance reporting and rehire eligibility tracking.",
        "Confirm whether terminated employees are intentionally excluded; if not, re-export the full workforce history.",
        None,   # dataset-level, no rows to show
    ),
    "status_high_pending": (
        "High Pending Status Rate",
        "Pending employees cannot be classified for payroll eligibility, headcount, or benefits enrollment.",
        "Update every Pending employee to the correct final status (Active, Inactive, or Terminated) before migration.",
        "Findings_Pending_Status",
    ),
    "age_uniformity": (
        "Age Data Issues",
        "Age-dependent benefits thresholds, compliance reporting, and retirement planning will be incorrect.",
        "Replace placeholder ages with actual employee date-of-birth data from the source system.",
        "Findings_Age_Data",
    ),
    "high_blank_rate": (
        "Missing Data",   # qualified with field name at runtime
        "Incomplete records reduce reporting accuracy and may cause downstream processing failures.",
        "Fill in missing values from the source system before using this data for any operational purpose.",
        "Findings_Completeness",
    ),
    "salary_outlier": (
        "Salary Outliers by Department",
        "Outlier salaries distort compensation analysis and may indicate data entry errors.",
        "Review each flagged salary against role and seniority benchmarks and correct any errors.",
        "Findings_Salary_Outliers",
    ),
    "missing_manager": (
        "Missing Manager Assignment",
        "The organisational hierarchy is incomplete - approval workflows and reporting structures will not function.",
        "Assign a valid manager ID to every active employee who lacks one.",
        "Findings_Missing_Mgr",
    ),
    # LOW
    "suspicious_round_salary": (
        "Suspicious Round Salaries",
        "Compensation data may be inaccurate, leading to incorrect pay or reporting.",
        "Verify the correct salary for each flagged employee with payroll records.",
        "Findings_Round_Salary",
    ),
    "combined_field": (
        "Combined Field",   # qualified with field name at runtime
        "The column cannot be mapped directly to any standard HRIS field without splitting.",
        "Split the column into two separate fields before loading into any target system.",
        None,   # affects the whole column, not individual rows
    ),
}

_DEFAULT_META = (
    None,   # issue_name → use check_name from JSON
    "This issue may affect data integrity and should be reviewed before migration or reporting.",
    "Review and correct all flagged records before using this data for any operational purpose.",
    None,   # no detail sheet
)


def _meta(check_key: str, check_name_fallback: str, field: str = "") -> tuple[str, str, str, str | None]:
    """Return (issue_name, business_impact, required_action, detail_sheet)."""
    base = _M.get(check_key, _DEFAULT_META)
    issue_name = base[0] or check_name_fallback
    # Qualify per-field issues
    if check_key in ("high_blank_rate", "combined_field") and field:
        issue_name = f"{base[0] or check_name_fallback} - {field.replace('_', ' ').title()}"
    return issue_name, base[1], base[2], base[3]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _read_and_normalize(
    file_path: Path, sheet_name: int | str = 0
) -> tuple[pd.DataFrame, dict[str, str], list[dict]]:
    df = ia._read_input(file_path, sheet_name=sheet_name)
    df = df.replace("", pd.NA)
    df.columns = [str(c).strip() for c in df.columns]
    try:
        from src.mapping import _apply_aliases  # noqa: PLC0415
        df = _apply_aliases(df)
    except Exception:
        pass
    col_map: dict[str, str] = {}
    for col in df.columns:
        norm = str(col).strip().lower().replace(" ", "_")
        col_map[col] = ia.ALIASES.get(norm, norm)
    df_norm_pre = df.rename(columns=col_map)
    _, _, row_annotations = ia.analyze_duplicate_canonical_fields(df_norm_pre, col_map)
    df_norm = ia._collapse_duplicate_columns(df_norm_pre)
    return df_norm, col_map, row_annotations


def _load_summary(run_dir: Path) -> dict:
    report_path = run_dir / "internal_audit_report.json"
    if not report_path.exists():
        raise FileNotFoundError(f"internal_audit_report.json not found in {run_dir}")
    return json.loads(report_path.read_text(encoding="utf-8"))


def _load_data_csv(run_dir: Path) -> pd.DataFrame | None:
    """Load internal_audit_data.csv and return only the audit-result rows."""
    path = run_dir / "internal_audit_data.csv"
    if not path.exists():
        return None
    raw = pd.read_csv(path)
    return raw[raw["check_name"].notna() & raw["severity"].notna()].copy()


def _humanize(col: str) -> str:
    display_map = {
        "worker_id": "Worker ID",
        "employee_id": "Employee ID",
        "first_name": "First Name",
        "last_name": "Last Name",
        "issue_name": "Issue Name",
        "issue_description": "Reason",
        "current_value": "Current Value",
        "expected_fix": "Recommended Action",
        "recommended_action": "Recommended Action",
        "required_action": "Required Action",
        "why_flagged": "Reason",
        "fix_needed": "Recommended Action",
        "row_number": "Row Number",
        "department": "Department",
        "district": "Department",
        "job_title": "Job Title",
        "title": "Job Title",
        "position_title": "Job Title",
        "position": "Job Title",
        "status": "Status",
        "worker_status": "Status",
        "pay_type": "Pay Type",
        "worker_type": "Pay Type",
        "salary": "Salary",
        "payrate": "Pay Rate",
        "salary_delta": "Salary Delta",
        "standard_hours": "Standard Hours",
        "hire_date": "Hire Date",
        "termination_date": "Termination Date",
        "records_affected": "Records Affected",
        "priority_rank": "Priority Rank",
        "what_to_do": "What To Do",
        "business_impact": "Business Impact",
        "gate_status": "Gate Status",
        "gate_message": "Gate Message",
        "source_file": "Source File",
        "detail_sheet": "Detail Sheet",
    }
    key = str(col).strip()
    lowered = key.lower().replace(" ", "_")
    if lowered in display_map:
        return display_map[lowered]
    return key.replace("_", " ").strip().title()


def _fmt(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.rename(columns={c: _humanize(c) for c in df.columns})


ROW_LEVEL_CORE_COLUMNS = [
    "Worker ID",
    "First Name",
    "Last Name",
    "Issue Name",
    "Severity",
    "Current Value",
    "Reason",
    "Recommended Action",
]

ROW_LEVEL_RELEVANT_COLUMNS = {
    "Fix_List_Detail": [
        "Department",
        "Job Title",
        "Status",
        "Pay Type",
        "Salary",
        "Pay Rate",
        "Salary Delta",
        "Standard Hours",
        "Hire Date",
        "Termination Date",
        "Row Number",
    ],
    "Findings_Missing_Salary": ["Salary", "Pay Rate", "Salary Delta"],
    "Findings_Pay_Type": ["Status", "Pay Type", "Salary", "Pay Rate"],
    "Findings_Comp_Type": ["Status", "Pay Type", "Salary", "Pay Rate"],
    "Findings_Comp_Conflict": ["Status", "Pay Type", "Salary", "Pay Rate", "Standard Hours"],
    "Findings_Std_Hours": ["Status", "Pay Type", "Pay Rate", "Standard Hours"],
    "Findings_Hourly_Rate": ["Status", "Pay Type", "Pay Rate", "Department", "Job Title"],
    "Findings_Impl_Salary": ["Status", "Pay Type", "Salary", "Department", "Job Title"],
    "Findings_Salary_Defaults": ["Salary", "Pay Rate", "Salary Delta"],
    "Findings_Round_Salary": ["Salary", "Pay Rate", "Salary Delta"],
    "Findings_Pay_Equity": ["Department", "Salary", "Salary Delta"],
    "Findings_Salary_Outliers": ["Department", "Salary", "Salary Delta"],
    "Findings_Active_Term": ["Status", "Termination Date"],
    "Findings_Pending_Status": ["Status", "Termination Date"],
    "Findings_Invalid_Dates": ["Hire Date", "Termination Date"],
    "Findings_Impossible_Dates": ["Hire Date", "Termination Date"],
}


def _reorder_columns(df: pd.DataFrame, first_columns: list[str]) -> pd.DataFrame:
    if df.empty:
        return df
    ordered = [col for col in first_columns if col in df.columns]
    remaining = [col for col in df.columns if col not in ordered]
    return df[[*ordered, *remaining]]


def _trim_row_level_columns(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    if df.empty:
        return df
    allowed = [col for col in ROW_LEVEL_CORE_COLUMNS if col in df.columns]
    relevant = [col for col in ROW_LEVEL_RELEVANT_COLUMNS.get(sheet_name, []) if col in df.columns]
    remaining = [col for col in df.columns if col not in set(allowed + relevant)]
    return df[[*allowed, *relevant, *remaining]]


def _format_row_level_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    formatted = _fmt(df)
    formatted = _trim_row_level_columns(formatted, sheet_name)
    return _reorder_columns(formatted, ROW_LEVEL_CORE_COLUMNS)


def _safe(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "nat", "none", "<na>") else s


def _records_affected(finding: dict) -> int | str:
    check_key = str(finding.get("check_key", ""))
    count = int(finding.get("count", finding.get("row_count", 0)) or 0)
    if check_key in DATASET_LEVEL_CHECKS:
        return "Dataset Level"
    return count


def _ordered_findings(summary: dict) -> list[dict]:
    """
    Return findings sorted CRITICAL→LOW then by count descending.
    Source of truth: findings_for_pdf or findings from JSON.
    """
    findings = summary.get("findings_for_pdf") or summary.get("findings") or []
    def _sort_key(f: dict) -> tuple[int, int]:
        sev = str(f.get("severity", "")).upper()
        cnt = int(f.get("count", f.get("row_count", 0)) or 0)
        return (SEVERITY_PRIORITY.get(sev, 9), -cnt)
    return sorted(findings, key=_sort_key)


# ── Sheet 1: Executive_Summary ────────────────────────────────────────────────

def _executive_summary_sheet(summary: dict) -> pd.DataFrame:
    """
    One row per finding, ordered by severity then count.
    Mirrors the PDF findings-by-severity section.
    Uses human-readable language throughout.
    """
    rows = [
        {
            "Severity": "",
            "Issue Name": "Start here",
            "Records Affected": "",
            "Description": "Fix CRITICAL issues first",
            "Business Impact": "",
            "Required Action": "Use Fix_List for priority-ranked issues",
        },
        {
            "Severity": "",
            "Issue Name": "Start here",
            "Records Affected": "",
            "Description": "Use Fix_List_Detail to see affected employees",
            "Business Impact": "",
            "Required Action": "Use correction CSV files to apply fixes",
        },
        {
            "Severity": "CRITICAL",
            "Issue Name": "Priority",
            "Records Affected": "",
            "Description": "Must fix immediately",
            "Business Impact": "",
            "Required Action": "",
        },
        {
            "Severity": "HIGH",
            "Issue Name": "Priority",
            "Records Affected": "",
            "Description": "Fix before migration",
            "Business Impact": "",
            "Required Action": "",
        },
        {
            "Severity": "MEDIUM",
            "Issue Name": "Priority",
            "Records Affected": "",
            "Description": "Review recommended",
            "Business Impact": "",
            "Required Action": "",
        },
        {
            "Severity": "LOW",
            "Issue Name": "Priority",
            "Records Affected": "",
            "Description": "Informational",
            "Business Impact": "",
            "Required Action": "",
        },
        {
            "Severity": "CRITICAL",
            "Issue Name": "Implausible pay findings",
            "Records Affected": "",
            "Description": "Clearly invalid compensation value such as zero or negative pay",
            "Business Impact": "",
            "Required Action": "",
        },
        {
            "Severity": "HIGH",
            "Issue Name": "Implausible pay findings",
            "Records Affected": "",
            "Description": "Unusual compensation value that should be reviewed before payroll or migration",
            "Business Impact": "",
            "Required Action": "",
        },
    ]
    for finding in _ordered_findings(summary):
        check_key  = str(finding.get("check_key", ""))
        check_name = str(finding.get("check_name", ""))
        sev        = str(finding.get("severity", "")).upper()
        field      = str(finding.get("field", ""))
        ra         = _records_affected(finding)
        description = str(finding.get("description", ""))

        issue_name, business_impact, required_action, _ = _meta(check_key, check_name, field)

        rows.append({
            "Severity":         sev,
            "Issue Name":       issue_name,
            "Records Affected": ra,
            "Description":      description,
            "Business Impact":  business_impact,
            "Required Action":  required_action,
        })

    if not rows:
        return pd.DataFrame([{"Note": "No issues found - dataset passed all checks."}])
    return pd.DataFrame(rows)


# ── Sheet 2: Fix_List ─────────────────────────────────────────────────────────

def _fix_list_sheet(summary: dict) -> pd.DataFrame:
    """
    All issues sorted by Priority Rank then count.
    Columns designed for HR action: what's wrong, how serious, what to do, blocking?
    """
    rows = []
    for finding in _ordered_findings(summary):
        check_key  = str(finding.get("check_key", ""))
        check_name = str(finding.get("check_name", ""))
        sev        = str(finding.get("severity", "")).upper()
        field      = str(finding.get("field", ""))
        ra         = _records_affected(finding)

        issue_name, _, required_action, _ = _meta(check_key, check_name, field)

        rows.append({
            "Issue Name":       issue_name,
            "Severity":         sev,
            "Records Affected": ra,
            "Priority Rank":    SEVERITY_PRIORITY.get(sev, 9),
            "What To Do":       required_action,
            "Blocking?":        "Yes" if sev in BLOCKING_SEVERITIES else "No",
        })

    if not rows:
        return pd.DataFrame([{"Note": "No issues found - dataset passed all checks."}])
    return pd.DataFrame(rows)


# ── Sheet 3: Fix_List_Detail ─────────────────────────────────────────────────

FIX_LIST_DETAIL_COLUMNS = [
    "Worker ID",
    "First Name",
    "Last Name",
    "Issue Name",
    "Severity",
    "Current Value",
    "Reason",
    "Recommended Action",
    "Pay Type",
    "Salary",
    "Pay Rate",
    "Salary Delta",
    "Standard Hours",
    "Status",
    "Hire Date",
    "Termination Date",
    "Department",
    "Row Number",
]

_RAW_EXECUTION_COLUMNS = [
    "Worker ID",
    "First Name",
    "Last Name",
    "Issue Name",
    "Severity",
    "Current Value",
    "Reason",
    "Recommended Action",
    "Department",
    "Status",
    "Pay Type",
    "Salary",
    "Pay Rate",
    "Salary Delta",
    "Standard Hours",
    "Hire Date",
    "Termination Date",
    "Row Number",
]


def _row_number_from_index(orig_idx: object) -> str:
    try:
        return str(int(orig_idx) + 2)
    except Exception:
        return ""


def _lookup_value(source_row: pd.Series | None, sample_row: dict | None, *keys: str) -> str:
    for key in keys:
        if source_row is not None and key in source_row.index:
            value = _safe(source_row.get(key, ""))
            if value:
                return value
        if sample_row is not None:
            value = _safe(sample_row.get(key, ""))
            if value:
                return value
    return ""


def _split_name_parts(sample_row: dict | None) -> tuple[str, str]:
    full_name = _safe((sample_row or {}).get("name", ""))
    if not full_name:
        return "", ""
    parts = full_name.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _source_row_from_sample(df: pd.DataFrame, sample_row: dict) -> pd.Series | None:
    row_number = sample_row.get("row_number")
    if row_number not in (None, ""):
        try:
            pos = int(row_number) - 2
        except Exception:
            pos = None
        if pos is not None and 0 <= pos < len(df):
            return df.iloc[pos]

    worker_id = _safe(sample_row.get("worker_id", ""))
    if worker_id and "worker_id" in df.columns:
        matches = df[df["worker_id"].astype(str).str.strip() == worker_id]
        if len(matches) == 1:
            return matches.iloc[0]
    return None


def _execution_why_flagged(
    check_key: str,
    issue_name: str,
    finding: dict,
    source_row: pd.Series | None,
    sample_row: dict | None,
) -> str:
    field = str(finding.get("field", "")).strip()
    field_label = _humanize(field) if field else ""

    if check_key == "duplicate_email":
        return "Email address appears on more than one employee record."
    if check_key == "active_zero_salary":
        return "Active employee has a missing, zero, or invalid salary/payrate."
    if check_key == "pay_type_missing_or_invalid":
        return "Compensation is present but pay type is blank or invalid."
    if check_key == "compensation_type_mismatch":
        return "Pay type does not match the required compensation field for this worker."
    if check_key == "comp_dual_value_conflict":
        return "Both salary and pay rate are populated and conflict with the worker pay type."
    if check_key == "missing_standard_hours_hourly":
        return "Hourly worker has pay rate present but standard hours are missing."
    if check_key == "phone_invalid":
        return "Phone number is invalid or impossible and should be corrected."
    if check_key == "status_high_pending":
        return "Worker status is Pending and needs a final employment status."
    if check_key == "duplicate_name_different_id":
        return "This employee name appears with different Worker IDs and needs review."
    if check_key == "high_blank_rate" and field_label:
        return f"{field_label} is blank on this record."
    if check_key == "missing_required_identity":
        missing = []
        for col, label in (("worker_id", "Worker ID"), ("first_name", "First Name"), ("last_name", "Last Name")):
            if _lookup_value(source_row, sample_row, col) == "":
                missing.append(label)
        if missing:
            return f"Required identity fields are missing: {', '.join(missing)}."
    return _safe(finding.get("description", "")) or f"{issue_name} requires review."


def _execution_current_value(
    check_key: str,
    finding: dict,
    source_row: pd.Series | None,
    sample_row: dict | None,
) -> str:
    if check_key == "duplicate_email":
        return _lookup_value(source_row, sample_row, "email")
    if check_key == "active_zero_salary":
        salary = _lookup_value(source_row, sample_row, "salary")
        payrate = _lookup_value(source_row, sample_row, "payrate")
        if salary and payrate:
            return f"Salary: {salary} | Pay Rate: {payrate}"
        if salary:
            return f"Salary: {salary}"
        if payrate:
            return f"Pay Rate: {payrate}"
        return "Missing compensation value"
    if check_key == "pay_type_missing_or_invalid":
        return _lookup_value(source_row, sample_row, "pay_type", "worker_type") or "Missing pay type"
    if check_key in {"compensation_type_mismatch", "comp_dual_value_conflict"}:
        return (
            f"Pay Type: {_lookup_value(source_row, sample_row, 'pay_type', 'worker_type')} | "
            f"Salary: {_lookup_value(source_row, sample_row, 'salary')} | "
            f"Pay Rate: {_lookup_value(source_row, sample_row, 'payrate')}"
        )
    if check_key == "missing_standard_hours_hourly":
        return (
            f"Pay Rate: {_lookup_value(source_row, sample_row, 'payrate')} | "
            f"Standard Hours: {_lookup_value(source_row, sample_row, 'standard_hours') or 'Missing'}"
        )
    if check_key == "phone_invalid":
        return _lookup_value(source_row, sample_row, "phone")
    if check_key == "status_high_pending":
        return _lookup_value(source_row, sample_row, "worker_status", "status")
    if check_key == "duplicate_name_different_id":
        worker_id = _lookup_value(source_row, sample_row, "worker_id")
        full_name = _safe((sample_row or {}).get("name", ""))
        if full_name and worker_id:
            return f"{full_name} | Worker ID: {worker_id}"
        return full_name or worker_id
    if check_key == "high_blank_rate":
        field = str(finding.get("field", "")).strip()
        return f"Blank {field.replace('_', ' ')}".strip()
    field = str(finding.get("field", "")).strip()
    if field:
        return _lookup_value(source_row, sample_row, field)
    return ""


def _execution_fix_needed(check_key: str, finding: dict, required_action: str) -> str:
    if check_key == "duplicate_email":
        return "Assign a unique, valid email address to this employee."
    if check_key == "active_zero_salary":
        return "Enter a valid positive salary or payrate before payroll processing."
    if check_key == "pay_type_missing_or_invalid":
        return "Populate a valid pay type from the controlled allowed list."
    if check_key == "compensation_type_mismatch":
        return "Populate the compensation field required by the worker pay type."
    if check_key == "comp_dual_value_conflict":
        return "Review the worker record and keep only the compensation field that matches the intended pay type."
    if check_key == "missing_standard_hours_hourly":
        return "Populate standard hours for each hourly worker."
    if check_key == "phone_invalid":
        return "Correct the phone number from the source record or a verified employee record."
    if check_key == "status_high_pending":
        return "Update the worker to the correct final status."
    if check_key == "duplicate_name_different_id":
        return "Confirm whether this is a duplicate person or different employees with the same name."
    if check_key == "high_blank_rate":
        field = str(finding.get("field", "")).strip()
        field_label = _humanize(field) if field else "missing value"
        return f"Populate {field_label} from the source system."
    return _safe(finding.get("recommended_action", "")) or required_action


def _execution_row(
    issue_name: str,
    severity: str,
    reason: str,
    current_value: str,
    recommended_action: str,
    source_row: pd.Series | None = None,
    sample_row: dict | None = None,
    row_number: str = "",
) -> dict:
    sample_first, sample_last = _split_name_parts(sample_row)
    result = {
        "Worker ID": _lookup_value(source_row, sample_row, "worker_id", "employee_id"),
        "First Name": _lookup_value(source_row, sample_row, "first_name") or sample_first,
        "Last Name": _lookup_value(source_row, sample_row, "last_name") or sample_last,
        "Issue Name": issue_name,
        "Severity": severity,
        "Current Value": current_value,
        "Reason": reason,
        "Recommended Action": recommended_action,
        "Pay Type": _lookup_value(source_row, sample_row, "pay_type", "worker_type"),
        "Salary": _lookup_value(source_row, sample_row, "salary"),
        "Pay Rate": _lookup_value(source_row, sample_row, "payrate"),
        "Salary Delta": _lookup_value(source_row, sample_row, "salary_delta"),
        "Standard Hours": _lookup_value(source_row, sample_row, "standard_hours"),
        "Status": _lookup_value(source_row, sample_row, "worker_status", "status"),
        "Hire Date": _lookup_value(source_row, sample_row, "hire_date", "start_date", "date_hired"),
        "Termination Date": _lookup_value(source_row, sample_row, "termination_date", "term_date", "end_date"),
        "Department": _lookup_value(source_row, sample_row, "department", "district"),
        "Job Title": _lookup_value(source_row, sample_row, "job_title", "title", "position_title", "position"),
        "Row Number": row_number or _lookup_value(source_row, sample_row, "row_number"),
    }
    return result


def _trim_optional_execution_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    return _format_row_level_sheet(frame, "Fix_List_Detail")


def _detail_extra_value(source_row: pd.Series | None, sample_row: dict | None, column: str) -> str:
    lookup_map = {
        "Status": ("worker_status", "status"),
        "Pay Type": ("pay_type", "worker_type"),
        "Salary": ("salary",),
        "Pay Rate": ("payrate",),
        "Salary Delta": ("salary_delta",),
        "Standard Hours": ("standard_hours",),
        "Hire Date": ("hire_date", "start_date", "date_hired"),
        "Termination Date": ("termination_date", "term_date", "end_date"),
        "Department": ("department", "district"),
        "Job Title": ("job_title", "title", "position_title", "position"),
    }
    return _lookup_value(source_row, sample_row, *lookup_map.get(column, ()))


DETAIL_COLUMNS = [
    "Worker ID",
    "First Name",
    "Last Name",
    "Issue Name",
    "Severity",
    "Current Value",
    "Reason",
    "Recommended Action",
    "Status",
    "Salary",
    "Pay Rate",
    "Salary Delta",
    "Hire Date",
    "Termination Date",
    "Department",
]


def _detail_row(
    issue_name: str,
    severity: str,
    current_value: str,
    reason: str,
    recommended_action: str,
    source_row: pd.Series | None = None,
    sample_row: dict | None = None,
    extra_columns: list[str] | None = None,
) -> dict:
    sample_first, sample_last = _split_name_parts(sample_row)
    row = {
        "Worker ID": _lookup_value(source_row, sample_row, "worker_id", "employee_id"),
        "First Name": _lookup_value(source_row, sample_row, "first_name") or sample_first,
        "Last Name": _lookup_value(source_row, sample_row, "last_name") or sample_last,
        "Issue Name": issue_name,
        "Severity": severity,
        "Current Value": current_value,
        "Reason": reason,
        "Recommended Action": recommended_action,
    }
    for column in extra_columns or []:
        row[column] = _detail_extra_value(source_row, sample_row, column)
    return row


def _format_detail_sheet(frame: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    if frame.empty:
        return frame
    return _format_row_level_sheet(frame, sheet_name)


def _build_fix_list_detail_from_samples(df: pd.DataFrame, summary: dict) -> pd.DataFrame:
    rows: list[dict] = []
    for finding in _ordered_findings(summary):
        check_key = str(finding.get("check_key", ""))
        check_name = str(finding.get("check_name", ""))
        severity = str(finding.get("severity", "")).upper()
        field = str(finding.get("field", ""))
        sample_rows = finding.get("sample_rows") or []
        if not sample_rows:
            continue

        issue_name, _, required_action, detail_sheet = _meta(check_key, check_name, field)
        if not detail_sheet or check_key in _FULL_DETAIL_BUILDERS:
            continue

        for sample_row in sample_rows:
            source_row = _source_row_from_sample(df, sample_row)
            rows.append(_execution_row(
                issue_name=issue_name,
                severity=severity,
                reason=_execution_why_flagged(check_key, issue_name, finding, source_row, sample_row),
                current_value=_execution_current_value(check_key, finding, source_row, sample_row),
                recommended_action=_execution_fix_needed(check_key, finding, required_action),
                source_row=source_row,
                sample_row=sample_row,
            ))

    if not rows:
        return pd.DataFrame()
    return _trim_optional_execution_columns(pd.DataFrame(rows, columns=FIX_LIST_DETAIL_COLUMNS))


def _build_fix_list_detail_full_dup_worker_id(
    df: pd.DataFrame,
    row_annotations: list[dict],
) -> pd.DataFrame:
    rows: list[dict] = []

    for idx, anns in enumerate(row_annotations):
        det = (anns or {}).get("worker_id")
        if not det or det.get("duplicate_classification") != "duplicate_conflicting_values":
            continue
        try:
            source_row = df.iloc[idx]
        except IndexError:
            continue
        rows.append(_execution_row(
            issue_name="Duplicate Worker ID (Source Column Conflict)",
            severity="CRITICAL",
            reason="Two source columns map to Worker ID with conflicting values.",
            current_value=_safe(det.get("duplicate_values", "")),
            recommended_action="Choose the authoritative Worker ID value and remove the conflicting one.",
            source_row=source_row,
            row_number=_row_number_from_index(idx),
        ))

    if "worker_id" in df.columns:
        series = df["worker_id"].astype(str).str.strip()
        nonblank = series[(series != "") & (series.str.lower() != "nan")]
        dup_idx = nonblank[nonblank.duplicated(keep=False)].index
        for orig_idx in dup_idx:
            source_row = df.loc[orig_idx]
            rows.append(_execution_row(
                issue_name="Duplicate Worker ID",
                severity="CRITICAL",
                reason="Worker ID appears on more than one employee record.",
                current_value=_safe(source_row.get("worker_id", "")),
                recommended_action="Assign a unique Worker ID to each affected employee.",
                source_row=source_row,
                row_number=_row_number_from_index(orig_idx),
            ))

    if not rows:
        return pd.DataFrame()
    return _trim_optional_execution_columns(pd.DataFrame(rows, columns=FIX_LIST_DETAIL_COLUMNS))


def _build_fix_list_detail_full_salary(df: pd.DataFrame) -> pd.DataFrame:
    status_col = ia._status_column(df)
    has_salary = "salary" in df.columns
    has_payrate = "payrate" in df.columns
    if not status_col or (not has_salary and not has_payrate):
        return pd.DataFrame()

    statuses = df[status_col].astype(str).str.strip().str.lower()
    sal_vals = pd.to_numeric(df["salary"], errors="coerce") if has_salary else pd.Series([float("nan")] * len(df), index=df.index)
    pay_vals = pd.to_numeric(df["payrate"], errors="coerce") if has_payrate else pd.Series([float("nan")] * len(df), index=df.index)
    sal_blank = ia._blank_mask(df["salary"]) if has_salary else pd.Series([True] * len(df), index=df.index)
    pay_blank = ia._blank_mask(df["payrate"]) if has_payrate else pd.Series([True] * len(df), index=df.index)
    effective = sal_vals.where(~sal_blank, pay_vals)
    comp_blank = sal_blank & pay_blank
    mask = (statuses == "active") & (comp_blank | (effective <= 0))

    rows = []
    for orig_idx in df.index[mask.fillna(False)]:
        source_row = df.loc[orig_idx]
        salary = _safe(source_row.get("salary", ""))
        payrate = _safe(source_row.get("payrate", ""))
        if salary and payrate:
            current_value = f"Salary: {salary} | Pay Rate: {payrate}"
        elif salary:
            current_value = f"Salary: {salary}"
        elif payrate:
            current_value = f"Pay Rate: {payrate}"
        else:
            current_value = "Missing compensation value"

        rows.append(_execution_row(
            issue_name="Missing or Invalid Salary",
            severity="CRITICAL",
            reason="Active employee has a missing, zero, or invalid salary/payrate.",
            current_value=current_value,
            recommended_action="Enter a valid positive salary or payrate before payroll processing.",
            source_row=source_row,
            row_number=_row_number_from_index(orig_idx),
        ))

    if not rows:
        return pd.DataFrame()
    return _trim_optional_execution_columns(pd.DataFrame(rows, columns=FIX_LIST_DETAIL_COLUMNS))


def _build_fix_list_detail_payroll_phase1(df: pd.DataFrame) -> pd.DataFrame:
    status_col = ia._status_column(df)
    pay_type_col = ia._pay_type_column(df) if hasattr(ia, "_pay_type_column") else ia._first_present(df, ["pay_type", "worker_type"])
    standard_hours_col = ia._standard_hours_column(df) if hasattr(ia, "_standard_hours_column") else ia._first_present(df, ["standard_hours"])
    has_salary = "salary" in df.columns
    has_payrate = "payrate" in df.columns
    if not status_col or not pay_type_col or (not has_salary and not has_payrate):
        return pd.DataFrame()

    statuses = df[status_col].astype(str).str.strip().str.lower()
    salary_blank = ia._blank_mask(df["salary"]) if has_salary else pd.Series([True] * len(df), index=df.index)
    payrate_blank = ia._blank_mask(df["payrate"]) if has_payrate else pd.Series([True] * len(df), index=df.index)
    salary_num = pd.to_numeric(df["salary"], errors="coerce") if has_salary else pd.Series([float("nan")] * len(df), index=df.index)
    payrate_num = pd.to_numeric(df["payrate"], errors="coerce") if has_payrate else pd.Series([float("nan")] * len(df), index=df.index)
    salary_valid = ~salary_blank & salary_num.gt(0)
    payrate_valid = ~payrate_blank & payrate_num.gt(0)
    comp_present = ~salary_blank | ~payrate_blank
    standard_hours_blank = ia._blank_mask(df[standard_hours_col]) if standard_hours_col else pd.Series([True] * len(df), index=df.index)

    rows: list[dict] = []
    for orig_idx in df.index:
        source_row = df.loc[orig_idx]
        pay_class, invalid = ia._classify_pay_type(source_row.get(pay_type_col, "")) if hasattr(ia, "_classify_pay_type") else ("", False)
        pay_type_blank = ia._blank_mask(pd.Series([source_row.get(pay_type_col, "")])).iloc[0]
        is_active = statuses.at[orig_idx] == "active"

        if comp_present.at[orig_idx] and (invalid or pay_type_blank):
            rows.append(_execution_row(
                issue_name="Missing or Invalid Pay Type",
                severity="CRITICAL" if is_active else "HIGH",
                reason="Compensation is present but pay type is blank or invalid.",
                current_value=_safe(source_row.get(pay_type_col, "")) or "Missing pay type",
                recommended_action="Populate a valid pay type from the controlled allowed list.",
                source_row=source_row,
                row_number=_row_number_from_index(orig_idx),
            ))
            continue

        if not pay_class:
            continue

        if comp_present.at[orig_idx]:
            if pay_class == "hourly" and not payrate_valid.at[orig_idx]:
                rows.append(_execution_row(
                    issue_name="Compensation Type Mismatch",
                    severity="CRITICAL" if is_active else "HIGH",
                    reason="Hourly worker is missing a valid pay rate for the stated pay type.",
                    current_value=f"Pay Type: {_safe(source_row.get(pay_type_col, ''))} | Salary: {_safe(source_row.get('salary', ''))} | Pay Rate: {_safe(source_row.get('payrate', ''))}",
                    recommended_action="Populate a valid pay rate that matches the worker pay type.",
                    source_row=source_row,
                    row_number=_row_number_from_index(orig_idx),
                ))
            elif pay_class == "salaried" and not salary_valid.at[orig_idx]:
                rows.append(_execution_row(
                    issue_name="Compensation Type Mismatch",
                    severity="CRITICAL" if is_active else "HIGH",
                    reason="Salaried worker is missing a valid salary for the stated pay type.",
                    current_value=f"Pay Type: {_safe(source_row.get(pay_type_col, ''))} | Salary: {_safe(source_row.get('salary', ''))} | Pay Rate: {_safe(source_row.get('payrate', ''))}",
                    recommended_action="Populate a valid salary that matches the worker pay type.",
                    source_row=source_row,
                    row_number=_row_number_from_index(orig_idx),
                ))

        if has_salary and has_payrate and not salary_blank.at[orig_idx] and not payrate_blank.at[orig_idx]:
            rows.append(_execution_row(
                issue_name="Salary and Pay Rate Conflict",
                severity="HIGH" if is_active else "MEDIUM",
                reason="Both salary and pay rate are populated and conflict with the worker pay type.",
                current_value=f"Pay Type: {_safe(source_row.get(pay_type_col, ''))} | Salary: {_safe(source_row.get('salary', ''))} | Pay Rate: {_safe(source_row.get('payrate', ''))}",
                recommended_action="Review the worker record and keep only the compensation field that matches the intended pay type.",
                source_row=source_row,
                row_number=_row_number_from_index(orig_idx),
            ))

        if standard_hours_col and pay_class == "hourly" and not payrate_blank.at[orig_idx] and standard_hours_blank.at[orig_idx]:
            rows.append(_execution_row(
                issue_name="Missing Standard Hours for Hourly Worker",
                severity="HIGH" if is_active else "MEDIUM",
                reason="Hourly worker has pay rate present but standard hours are missing.",
                current_value=f"Pay Rate: {_safe(source_row.get('payrate', ''))} | Standard Hours: Missing",
                recommended_action="Populate standard hours for each hourly worker.",
                source_row=source_row,
                row_number=_row_number_from_index(orig_idx),
            ))

    if not rows:
        return pd.DataFrame()
    return _trim_optional_execution_columns(pd.DataFrame(rows, columns=FIX_LIST_DETAIL_COLUMNS))


def _build_fix_list_detail_full_invalid_dates(df: pd.DataFrame) -> pd.DataFrame:
    hire_col = ia._first_present(df, ["hire_date", "start_date", "date_hired"])
    term_col = ia._first_present(df, ["termination_date", "term_date", "end_date"])
    if not hire_col:
        return pd.DataFrame()

    today = pd.Timestamp.now().normalize().tz_localize(None)
    hire_dates = ia._date_series(df, hire_col)
    rows = []

    future_mask = hire_dates > today
    for orig_idx in df.index[future_mask.fillna(False)]:
        source_row = df.loc[orig_idx]
        rows.append(_execution_row(
            issue_name="Invalid Dates",
            severity="CRITICAL",
            reason="Hire date is set in the future.",
            current_value=_safe(source_row.get(hire_col, "")),
            recommended_action="Correct the hire date to today or earlier.",
            source_row=source_row,
            row_number=_row_number_from_index(orig_idx),
        ))

    if term_col:
        term_dates = ia._date_series(df, term_col)
        tbh_mask = (term_dates < hire_dates) & term_dates.notna() & hire_dates.notna()
        for orig_idx in df.index[tbh_mask.fillna(False)]:
            source_row = df.loc[orig_idx]
            rows.append(_execution_row(
                issue_name="Invalid Dates",
                severity="CRITICAL",
                reason="Termination date is earlier than hire date.",
                current_value=f"Hire Date: {_safe(source_row.get(hire_col, ''))} | Termination Date: {_safe(source_row.get(term_col, ''))}",
                recommended_action="Update the dates so the termination date is on or after the hire date.",
                source_row=source_row,
                row_number=_row_number_from_index(orig_idx),
            ))

    if not rows:
        return pd.DataFrame()
    return _trim_optional_execution_columns(pd.DataFrame(rows, columns=FIX_LIST_DETAIL_COLUMNS))


def _build_fix_list_detail_full_active_term(df: pd.DataFrame) -> pd.DataFrame:
    status_col = ia._status_column(df)
    term_col = ia._first_present(df, ["termination_date", "term_date", "end_date"])
    if not status_col or not term_col:
        return pd.DataFrame()

    statuses = df[status_col].astype(str).str.strip().str.lower()
    term_blank = ia._blank_mask(df[term_col])
    mask = (statuses == "active") & ~term_blank

    rows = []
    for orig_idx in df.index[mask.fillna(False)]:
        source_row = df.loc[orig_idx]
        rows.append(_execution_row(
            issue_name="Active Employees with Termination Dates",
            severity="CRITICAL",
            reason="Employee is marked Active but also has a termination date.",
            current_value=f"Status: Active | Termination Date: {_safe(source_row.get(term_col, ''))}",
            recommended_action="Remove the termination date or change the worker status to Terminated.",
            source_row=source_row,
            row_number=_row_number_from_index(orig_idx),
        ))

    if not rows:
        return pd.DataFrame()
    return _trim_optional_execution_columns(pd.DataFrame(rows, columns=FIX_LIST_DETAIL_COLUMNS))


def _build_fix_list_detail_full_missing_id(df: pd.DataFrame) -> pd.DataFrame:
    required = [f for f in ["worker_id", "first_name", "last_name"] if f in df.columns]
    if not required:
        return pd.DataFrame()

    masks = [ia._blank_mask(df[f]) for f in required]
    combined = masks[0].copy()
    for mask in masks[1:]:
        combined = combined | mask

    labels = {"worker_id": "Worker ID", "first_name": "First Name", "last_name": "Last Name"}
    rows = []
    for orig_idx in df.index[combined.fillna(False)]:
        source_row = df.loc[orig_idx]
        missing = [labels.get(field, field) for field in required if _safe(source_row.get(field, "")) == ""]
        rows.append(_execution_row(
            issue_name="Missing Required Identity Fields",
            severity="CRITICAL",
            reason=f"Required identity fields are missing: {', '.join(missing)}.",
            current_value="Missing: " + ", ".join(missing),
            recommended_action="Populate Worker ID, First Name, and Last Name for this employee.",
            source_row=source_row,
            row_number=_row_number_from_index(orig_idx),
        ))

    if not rows:
        return pd.DataFrame()
    return _trim_optional_execution_columns(pd.DataFrame(rows, columns=FIX_LIST_DETAIL_COLUMNS))


def _fix_list_detail_sheet(df: pd.DataFrame, summary: dict, row_annotations: list[dict]) -> pd.DataFrame:
    frames = [
        _build_fix_list_detail_full_dup_worker_id(df, row_annotations),
        _build_fix_list_detail_full_salary(df),
        _build_fix_list_detail_payroll_phase1(df),
        _build_fix_list_detail_full_invalid_dates(df),
        _build_fix_list_detail_full_active_term(df),
        _build_fix_list_detail_full_missing_id(df),
        _build_fix_list_detail_from_samples(df, summary),
    ]

    nonempty = [frame for frame in frames if frame is not None and not frame.empty]
    if not nonempty:
        return pd.DataFrame([{
            "Worker ID": "",
            "First Name": "",
            "Last Name": "",
            "Issue Name": "No Row-Level Issues",
            "Severity": "",
            "Current Value": "",
            "Reason": "No actionable row-level detail was available for this workbook.",
            "Recommended Action": "",
        }])

    combined = pd.concat(nonempty, ignore_index=True, sort=False)
    combined["__severity_rank"] = combined["Severity"].map(SEVERITY_PRIORITY).fillna(99)
    combined["__worker_sort"] = combined["Worker ID"].map(_safe)
    combined["__issue_sort"] = combined["Issue Name"].map(_safe)
    combined["__row_sort"] = combined.get("Row Number", "").map(_safe) if "Row Number" in combined.columns else ""
    combined = combined.sort_values(
        by=["__severity_rank", "__issue_sort", "__worker_sort", "__row_sort"],
        kind="stable",
    ).drop(columns=["__severity_rank", "__worker_sort", "__issue_sort", "__row_sort"], errors="ignore")

    return _trim_optional_execution_columns(combined).reset_index(drop=True)


# ── Sheet 4: Findings_Detail (per issue) ─────────────────────────────────────

def _detail_note_row(msg: str) -> dict:
    return {
        "Worker ID": "Note",
        "First Name": "",
        "Last Name": "",
        "Issue Name": "",
        "Severity": "",
        "Current Value": "",
        "Reason": msg,
        "Recommended Action": "",
    }


def _detail_columns_for_sheet(sheet_name: str) -> list[str]:
    return [col for col in ROW_LEVEL_CORE_COLUMNS + ROW_LEVEL_RELEVANT_COLUMNS.get(sheet_name, [])]


def _build_sample_detail(
    data_csv_df: pd.DataFrame | None,
    check_name: str,
    finding: dict,
    sheet_name: str,
) -> pd.DataFrame:
    """Build a detail tab from internal_audit_data.csv sample rows."""
    if data_csv_df is None or data_csv_df.empty:
        total = int(finding.get("count", finding.get("row_count", 0)) or 0)
        return _format_detail_sheet(pd.DataFrame([_detail_note_row(
            f"No row-level sample data available. This issue affects {total:,} records. Export internal_audit_data.csv for the full list."
        )]), sheet_name)

    rows = data_csv_df[data_csv_df["check_name"] == check_name].copy()
    if rows.empty:
        return pd.DataFrame()

    total = int(finding.get("count", finding.get("row_count", 0)) or 0)
    detail = pd.DataFrame({
        "Worker ID": rows["employee_id"].map(_safe),
        "First Name": rows["first_name"].map(_safe),
        "Last Name": rows["last_name"].map(_safe),
        "Issue Name": rows["check_name"].map(_safe),
        "Severity": str(finding.get("severity", "")).upper(),
        "Current Value": rows["value_found"].map(_safe),
        "Reason": rows["issue_description"].map(_safe),
        "Recommended Action": rows["recommended_action"].map(_safe),
    }).reset_index(drop=True)

    if total > len(detail):
        note = _detail_note_row(
            f"Showing {len(detail)} sample records of {total:,} total. See internal_audit_data.csv for the complete list."
        )
        detail = pd.concat([pd.DataFrame([note]), detail], ignore_index=True)

    detail = detail.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(detail, sheet_name)


def _build_full_detail_dup_worker_id(
    df: pd.DataFrame,
    summary: dict,
    row_annotations: list[dict],
) -> pd.DataFrame:
    sheet_name = "Findings_Dup_WorkerID"
    frames = []

    # Canonical conflict rows
    for idx, anns in enumerate(row_annotations):
        det = (anns or {}).get("worker_id")
        if not det or det.get("duplicate_classification") != "duplicate_conflicting_values":
            continue
        try:
            row = df.iloc[idx]
        except IndexError:
            continue
        frames.append(_detail_row(
            issue_name="Duplicate Worker ID (Source Column Conflict)",
            severity="CRITICAL",
            current_value=_safe(det.get("duplicate_values", "")),
            reason="Two source columns map to Worker ID with conflicting values.",
            recommended_action="Decide which source column is authoritative and remove the other.",
            source_row=row,
        ))

    # Standard duplicate worker_id rows
    if "worker_id" in df.columns:
        series = df["worker_id"].astype(str).str.strip()
        nonblank = series[(series != "") & (series.str.lower() != "nan")]
        dup_idx = nonblank[nonblank.duplicated(keep=False)].index
        for orig_idx in dup_idx:
            row = df.loc[orig_idx]
            frames.append(_detail_row(
                issue_name="Duplicate Worker ID",
                severity="CRITICAL",
                current_value=_safe(row.get("worker_id", "")),
                reason="Multiple employees share this Worker ID.",
                recommended_action="Assign a unique Worker ID to each employee.",
                source_row=row,
            ))

    if not frames:
        return pd.DataFrame()
    result = pd.DataFrame(frames)
    if len(result) > DETAIL_ROW_CAP:
        note = _detail_note_row(f"Truncated to {DETAIL_ROW_CAP:,} rows. Full list in internal_audit_data.csv.")
        result = pd.concat([pd.DataFrame([note]), result.head(DETAIL_ROW_CAP)], ignore_index=True)
    result = result.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(result, sheet_name)


def _build_full_detail_salary(df: pd.DataFrame) -> pd.DataFrame:
    sheet_name = "Findings_Missing_Salary"
    status_col  = ia._status_column(df)
    has_salary  = "salary" in df.columns
    has_payrate = "payrate" in df.columns
    if not status_col or (not has_salary and not has_payrate):
        return pd.DataFrame()

    statuses = df[status_col].astype(str).str.strip().str.lower()
    sal_vals  = pd.to_numeric(df["salary"],  errors="coerce") if has_salary  else pd.Series([float("nan")] * len(df), index=df.index)
    pay_vals  = pd.to_numeric(df["payrate"], errors="coerce") if has_payrate else pd.Series([float("nan")] * len(df), index=df.index)
    sal_blank = ia._blank_mask(df["salary"])  if has_salary  else pd.Series([True] * len(df), index=df.index)
    pay_blank = ia._blank_mask(df["payrate"]) if has_payrate else pd.Series([True] * len(df), index=df.index)
    effective = sal_vals.where(~sal_blank, pay_vals)
    comp_blank = sal_blank & pay_blank
    mask = (statuses == "active") & (comp_blank | (effective <= 0))

    rows = []
    for orig_idx in df.index[mask.fillna(False)]:
        row = df.loc[orig_idx]
        sal = _safe(row.get("salary", ""))
        pay = _safe(row.get("payrate", ""))
        current = "Missing" if (not sal and not pay) else (f"Salary: {sal}" if sal else f"Pay Rate: {pay}")
        rows.append(_detail_row(
            issue_name="Missing or Invalid Salary",
            severity="CRITICAL",
            current_value=current,
            reason="Active employee has no valid salary or payrate.",
            recommended_action="Enter a valid positive salary or payrate.",
            source_row=row,
            extra_columns=["Salary", "Pay Rate", "Salary Delta"],
        ))

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    if len(result) > DETAIL_ROW_CAP:
        note = _detail_note_row(f"Truncated to {DETAIL_ROW_CAP:,} rows. Full list in internal_audit_data.csv.")
        result = pd.concat([pd.DataFrame([note]), result.head(DETAIL_ROW_CAP)], ignore_index=True)
    result = result.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(result, sheet_name)


def _build_full_detail_payroll_phase1(df: pd.DataFrame, issue_name: str, sheet_name: str) -> pd.DataFrame:
    status_col = ia._status_column(df)
    pay_type_col = ia._pay_type_column(df) if hasattr(ia, "_pay_type_column") else ia._first_present(df, ["pay_type", "worker_type"])
    standard_hours_col = ia._standard_hours_column(df) if hasattr(ia, "_standard_hours_column") else ia._first_present(df, ["standard_hours"])
    has_salary = "salary" in df.columns
    has_payrate = "payrate" in df.columns
    if not status_col or not pay_type_col or (not has_salary and not has_payrate):
        return pd.DataFrame()

    statuses = df[status_col].astype(str).str.strip().str.lower()
    salary_blank = ia._blank_mask(df["salary"]) if has_salary else pd.Series([True] * len(df), index=df.index)
    payrate_blank = ia._blank_mask(df["payrate"]) if has_payrate else pd.Series([True] * len(df), index=df.index)
    salary_num = pd.to_numeric(df["salary"], errors="coerce") if has_salary else pd.Series([float("nan")] * len(df), index=df.index)
    payrate_num = pd.to_numeric(df["payrate"], errors="coerce") if has_payrate else pd.Series([float("nan")] * len(df), index=df.index)
    salary_valid = ~salary_blank & salary_num.gt(0)
    payrate_valid = ~payrate_blank & payrate_num.gt(0)
    comp_present = ~salary_blank | ~payrate_blank
    standard_hours_blank = ia._blank_mask(df[standard_hours_col]) if standard_hours_col else pd.Series([True] * len(df), index=df.index)

    rows: list[dict] = []
    for orig_idx in df.index:
        row = df.loc[orig_idx]
        pay_class, invalid = ia._classify_pay_type(row.get(pay_type_col, "")) if hasattr(ia, "_classify_pay_type") else ("", False)
        pay_type_blank = ia._blank_mask(pd.Series([row.get(pay_type_col, "")])).iloc[0]
        is_active = statuses.at[orig_idx] == "active"

        if issue_name == "Missing or Invalid Pay Type":
            if comp_present.at[orig_idx] and (invalid or pay_type_blank):
                rows.append(_detail_row(
                    issue_name=issue_name,
                    severity="CRITICAL" if is_active else "HIGH",
                    current_value=_safe(row.get(pay_type_col, "")) or "Missing pay type",
                    reason="Compensation is present but pay type is blank or invalid.",
                    recommended_action="Populate a valid pay type from the controlled allowed list before payroll or migration.",
                    source_row=row,
                    extra_columns=["Status", "Pay Type", "Salary", "Pay Rate"],
                ))
        elif issue_name == "Compensation Type Mismatch":
            if not pay_class:
                continue
            if pay_class == "hourly" and comp_present.at[orig_idx] and not payrate_valid.at[orig_idx]:
                rows.append(_detail_row(
                    issue_name=issue_name,
                    severity="CRITICAL" if is_active else "HIGH",
                    current_value=f"Pay Type: {_safe(row.get(pay_type_col, ''))} | Salary: {_safe(row.get('salary', ''))} | Pay Rate: {_safe(row.get('payrate', ''))}",
                    reason="Hourly worker is missing a valid pay rate for the stated pay type.",
                    recommended_action="Populate a valid pay rate that matches the worker pay type before payroll.",
                    source_row=row,
                    extra_columns=["Status", "Pay Type", "Salary", "Pay Rate"],
                ))
            elif pay_class == "salaried" and comp_present.at[orig_idx] and not salary_valid.at[orig_idx]:
                rows.append(_detail_row(
                    issue_name=issue_name,
                    severity="CRITICAL" if is_active else "HIGH",
                    current_value=f"Pay Type: {_safe(row.get(pay_type_col, ''))} | Salary: {_safe(row.get('salary', ''))} | Pay Rate: {_safe(row.get('payrate', ''))}",
                    reason="Salaried worker is missing a valid salary for the stated pay type.",
                    recommended_action="Populate a valid salary that matches the worker pay type before payroll.",
                    source_row=row,
                    extra_columns=["Status", "Pay Type", "Salary", "Pay Rate"],
                ))
        elif issue_name == "Salary and Pay Rate Conflict":
            if not pay_class or not has_salary or not has_payrate or salary_blank.at[orig_idx] or payrate_blank.at[orig_idx]:
                continue
            rows.append(_detail_row(
                issue_name=issue_name,
                severity="HIGH" if is_active else "MEDIUM",
                current_value=f"Pay Type: {_safe(row.get(pay_type_col, ''))} | Salary: {_safe(row.get('salary', ''))} | Pay Rate: {_safe(row.get('payrate', ''))}",
                reason="Both salary and pay rate are populated and conflict with the worker pay type.",
                recommended_action="Review the worker record and keep only the compensation field that matches the intended pay type.",
                source_row=row,
                extra_columns=["Status", "Pay Type", "Salary", "Pay Rate", "Standard Hours"],
            ))
        elif issue_name == "Missing Standard Hours for Hourly Worker":
            if not standard_hours_col or pay_class != "hourly" or payrate_blank.at[orig_idx] or not standard_hours_blank.at[orig_idx]:
                continue
            rows.append(_detail_row(
                issue_name=issue_name,
                severity="HIGH" if is_active else "MEDIUM",
                current_value=f"Pay Rate: {_safe(row.get('payrate', ''))} | Standard Hours: Missing",
                reason="Hourly worker has pay rate present but standard hours are missing.",
                recommended_action="Populate standard hours for each hourly worker before migration.",
                source_row=row,
                extra_columns=["Status", "Pay Type", "Pay Rate", "Standard Hours"],
            ))

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    if len(result) > DETAIL_ROW_CAP:
        note = _detail_note_row(f"Truncated to {DETAIL_ROW_CAP:,} rows. Full list in internal_audit_data.csv.")
        result = pd.concat([pd.DataFrame([note]), result.head(DETAIL_ROW_CAP)], ignore_index=True)
    result = result.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(result, sheet_name)


def _build_full_detail_payroll_phase2(df: pd.DataFrame, issue_name: str, sheet_name: str) -> pd.DataFrame:
    pay_type_col = ia._pay_type_column(df) if hasattr(ia, "_pay_type_column") else ia._first_present(df, ["pay_type", "worker_type"])
    if not pay_type_col:
        return pd.DataFrame()

    config = ia._load_config() if hasattr(ia, "_load_config") else {}
    hourly_min = float(config.get("hourly_payrate_min", ia.DEFAULT_HOURLY_PAYRATE_MIN))
    hourly_max = float(config.get("hourly_payrate_max", ia.DEFAULT_HOURLY_PAYRATE_MAX))
    salary_min = float(config.get("salaried_salary_min", ia.DEFAULT_SALARIED_SALARY_MIN))
    salary_max = float(config.get("salaried_salary_max", ia.DEFAULT_SALARIED_SALARY_MAX))
    pay_blank = ia._blank_mask(df["payrate"]) if "payrate" in df.columns else pd.Series([True] * len(df), index=df.index)
    sal_blank = ia._blank_mask(df["salary"]) if "salary" in df.columns else pd.Series([True] * len(df), index=df.index)
    pay_vals = pd.to_numeric(df["payrate"], errors="coerce") if "payrate" in df.columns else pd.Series([float("nan")] * len(df), index=df.index)
    sal_vals = pd.to_numeric(df["salary"], errors="coerce") if "salary" in df.columns else pd.Series([float("nan")] * len(df), index=df.index)

    rows: list[dict] = []
    for orig_idx in df.index:
        row = df.loc[orig_idx]
        pay_class, invalid = ia._classify_pay_type(row.get(pay_type_col, "")) if hasattr(ia, "_classify_pay_type") else ("", False)
        if invalid or not pay_class:
            continue

        if issue_name == "Implausible Hourly Pay Rate":
            if pay_class != "hourly" or pay_blank.at[orig_idx]:
                continue
            value = pay_vals.at[orig_idx]
            if pd.isna(value):
                continue
            if value <= 0:
                rows.append(_detail_row(
                    issue_name=issue_name,
                    severity="CRITICAL",
                    current_value=f"Pay Rate: {_safe(row.get('payrate', ''))}",
                    reason="Hourly worker has a zero or negative pay rate.",
                    recommended_action="Enter a valid positive pay rate before payroll.",
                    source_row=row,
                    extra_columns=["Status", "Pay Type", "Pay Rate", "Department", "Job Title"],
                ))
            elif value < hourly_min or value > hourly_max:
                rows.append(_detail_row(
                    issue_name=issue_name,
                    severity="HIGH",
                    current_value=f"Pay Rate: {_safe(row.get('payrate', ''))}",
                    reason=f"Hourly worker pay rate is outside the configured review range of {hourly_min:g} to {hourly_max:g}.",
                    recommended_action=f"Review the hourly pay rate and confirm it belongs within the configured range of {hourly_min:g} to {hourly_max:g}.",
                    source_row=row,
                    extra_columns=["Status", "Pay Type", "Pay Rate", "Department", "Job Title"],
                ))
        elif issue_name == "Implausible Annual Salary":
            if pay_class != "salaried" or sal_blank.at[orig_idx]:
                continue
            value = sal_vals.at[orig_idx]
            if pd.isna(value):
                continue
            if value <= 0:
                rows.append(_detail_row(
                    issue_name=issue_name,
                    severity="CRITICAL",
                    current_value=f"Salary: {_safe(row.get('salary', ''))}",
                    reason="Salaried worker has a zero or negative annual salary.",
                    recommended_action="Enter a valid positive salary before payroll.",
                    source_row=row,
                    extra_columns=["Status", "Pay Type", "Salary", "Department", "Job Title"],
                ))
            elif value < salary_min or value > salary_max:
                rows.append(_detail_row(
                    issue_name=issue_name,
                    severity="HIGH",
                    current_value=f"Salary: {_safe(row.get('salary', ''))}",
                    reason=f"Salaried worker salary is outside the configured review range of {salary_min:g} to {salary_max:g}.",
                    recommended_action=f"Review the salary and confirm it belongs within the configured range of {salary_min:g} to {salary_max:g}.",
                    source_row=row,
                    extra_columns=["Status", "Pay Type", "Salary", "Department", "Job Title"],
                ))

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    if len(result) > DETAIL_ROW_CAP:
        note = _detail_note_row(f"Truncated to {DETAIL_ROW_CAP:,} rows. Full list in internal_audit_data.csv.")
        result = pd.concat([pd.DataFrame([note]), result.head(DETAIL_ROW_CAP)], ignore_index=True)
    result = result.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(result, sheet_name)


def _build_full_detail_invalid_dates(df: pd.DataFrame) -> pd.DataFrame:
    sheet_name = "Findings_Invalid_Dates"
    hire_col = ia._first_present(df, ["hire_date", "start_date", "date_hired"])
    term_col  = ia._first_present(df, ["termination_date", "term_date", "end_date"])
    if not hire_col:
        return pd.DataFrame()

    today = pd.Timestamp.now().normalize().tz_localize(None)
    hire_dates = ia._date_series(df, hire_col)
    rows = []

    future_mask = hire_dates > today
    for orig_idx in df.index[future_mask.fillna(False)]:
        row = df.loc[orig_idx]
        rows.append(_detail_row(
            issue_name="Invalid Dates",
            severity="CRITICAL",
            current_value=_safe(row.get(hire_col, "")),
            reason="Hire date is set in the future.",
            recommended_action="Correct the hire date to today or earlier.",
            source_row=row,
            extra_columns=["Hire Date", "Termination Date"],
        ))

    if term_col:
        term_dates = ia._date_series(df, term_col)
        tbh_mask = (term_dates < hire_dates) & term_dates.notna() & hire_dates.notna()
        for orig_idx in df.index[tbh_mask.fillna(False)]:
            row = df.loc[orig_idx]
            rows.append(_detail_row(
                issue_name="Invalid Dates",
                severity="CRITICAL",
                current_value=f"Hire Date: {_safe(row.get(hire_col, ''))} | Termination Date: {_safe(row.get(term_col, ''))}",
                reason="Termination date is before hire date.",
                recommended_action="Ensure termination date is on or after the hire date.",
                source_row=row,
                extra_columns=["Hire Date", "Termination Date"],
            ))

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    if len(result) > DETAIL_ROW_CAP:
        note = _detail_note_row(f"Truncated to {DETAIL_ROW_CAP:,} rows.")
        result = pd.concat([pd.DataFrame([note]), result.head(DETAIL_ROW_CAP)], ignore_index=True)
    result = result.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(result, sheet_name)


def _build_full_detail_active_term(df: pd.DataFrame) -> pd.DataFrame:
    sheet_name = "Findings_Active_Term"
    status_col = ia._status_column(df)
    term_col   = ia._first_present(df, ["termination_date", "term_date", "end_date"])
    if not status_col or not term_col:
        return pd.DataFrame()

    statuses  = df[status_col].astype(str).str.strip().str.lower()
    term_blank = ia._blank_mask(df[term_col])
    mask = (statuses == "active") & ~term_blank

    rows = []
    for orig_idx in df.index[mask.fillna(False)]:
        row = df.loc[orig_idx]
        rows.append(_detail_row(
            issue_name="Active Employees with Termination Dates",
            severity="CRITICAL",
            current_value=f"Status: Active | Termination Date: {_safe(row.get(term_col, ''))}",
            reason="Employee is marked Active but also has a termination date.",
            recommended_action="Remove the termination date, or change status to Terminated.",
            source_row=row,
            extra_columns=["Status", "Termination Date"],
        ))

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    if len(result) > DETAIL_ROW_CAP:
        note = _detail_note_row(f"Truncated to {DETAIL_ROW_CAP:,} rows.")
        result = pd.concat([pd.DataFrame([note]), result.head(DETAIL_ROW_CAP)], ignore_index=True)
    result = result.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(result, sheet_name)


def _build_full_detail_missing_id(df: pd.DataFrame) -> pd.DataFrame:
    sheet_name = "Findings_Missing_ID"
    required = [f for f in ["worker_id", "first_name", "last_name"] if f in df.columns]
    if not required:
        return pd.DataFrame()

    masks = [ia._blank_mask(df[f]) for f in required]
    combined = masks[0].copy()
    for m in masks[1:]:
        combined = combined | m

    _labels = {"worker_id": "Worker ID", "first_name": "First Name", "last_name": "Last Name"}
    rows = []
    for orig_idx in df.index[combined.fillna(False)]:
        row = df.loc[orig_idx]
        missing = [_labels.get(f, f) for f in required if _safe(row.get(f, "")) == ""]
        rows.append(_detail_row(
            issue_name="Missing Required Identity Fields",
            severity="CRITICAL",
            current_value="Missing: " + ", ".join(missing),
            reason=f"Required fields are blank: {', '.join(missing)}.",
            recommended_action="Provide Worker ID, First Name, and Last Name for every employee.",
            source_row=row,
        ))

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    if len(result) > DETAIL_ROW_CAP:
        note = _detail_note_row(f"Truncated to {DETAIL_ROW_CAP:,} rows.")
        result = pd.concat([pd.DataFrame([note]), result.head(DETAIL_ROW_CAP)], ignore_index=True)
    result = result.reindex(columns=_detail_columns_for_sheet(sheet_name), fill_value="")
    return _format_detail_sheet(result, sheet_name)


# Full-detail dispatch table: check_key → builder function (takes df, summary, row_annotations)
_FULL_DETAIL_BUILDERS = {
    "duplicate_worker_id":                    lambda df, s, ra: _build_full_detail_dup_worker_id(df, s, ra),
    "duplicate_canonical_worker_id_conflict": lambda df, s, ra: _build_full_detail_dup_worker_id(df, s, ra),
    "active_zero_salary":                     lambda df, s, ra: _build_full_detail_salary(df),
    "pay_type_missing_or_invalid":            lambda df, s, ra: _build_full_detail_payroll_phase1(df, "Missing or Invalid Pay Type", "Findings_Pay_Type"),
    "compensation_type_mismatch":             lambda df, s, ra: _build_full_detail_payroll_phase1(df, "Compensation Type Mismatch", "Findings_Comp_Type"),
    "hourly_implausible_payrate":             lambda df, s, ra: _build_full_detail_payroll_phase2(df, "Implausible Hourly Pay Rate", "Findings_Hourly_Rate"),
    "salaried_implausible_salary":            lambda df, s, ra: _build_full_detail_payroll_phase2(df, "Implausible Annual Salary", "Findings_Impl_Salary"),
    "comp_dual_value_conflict":               lambda df, s, ra: _build_full_detail_payroll_phase1(df, "Salary and Pay Rate Conflict", "Findings_Comp_Conflict"),
    "missing_standard_hours_hourly":          lambda df, s, ra: _build_full_detail_payroll_phase1(df, "Missing Standard Hours for Hourly Worker", "Findings_Std_Hours"),
    "invalid_date_logic":                     lambda df, s, ra: _build_full_detail_invalid_dates(df),
    "active_with_termination_date":           lambda df, s, ra: _build_full_detail_active_term(df),
    "missing_required_identity":              lambda df, s, ra: _build_full_detail_missing_id(df),
}


def _build_findings_detail_sheets(
    df: pd.DataFrame,
    summary: dict,
    row_annotations: list[dict],
    run_dir: Path,
) -> dict[str, pd.DataFrame]:
    """
    Returns { sheet_name: DataFrame } for all per-issue detail tabs.

    CRITICAL checks → full row expansion from source data.
    HIGH/MEDIUM checks → sample rows from internal_audit_data.csv + count note.
    Dataset-level checks (status_no_terminated) → no tab.
    """
    data_csv = _load_data_csv(run_dir)
    findings  = _ordered_findings(summary)

    result: dict[str, pd.DataFrame] = {}
    populated: set[str] = set()  # avoid building a sheet twice (e.g., dup_worker_id + canonical)

    for finding in findings:
        check_key  = str(finding.get("check_key", ""))
        check_name = str(finding.get("check_name", ""))
        field      = str(finding.get("field", ""))
        count      = int(finding.get("count", finding.get("row_count", 0)) or 0)

        if count == 0:
            continue

        _, _, _, detail_sheet = _meta(check_key, check_name, field)

        if not detail_sheet:
            continue  # dataset-level; no tab
        if detail_sheet in populated:
            continue

        if check_key in _FULL_DETAIL_BUILDERS:
            frame = _FULL_DETAIL_BUILDERS[check_key](df, summary, row_annotations)
        else:
            frame = _build_sample_detail(data_csv, check_name, finding, detail_sheet)

        if frame is not None and not frame.empty:
            result[detail_sheet] = frame
            populated.add(detail_sheet)

    return result


# ── Sheet 5: Findings_Index ───────────────────────────────────────────────────

def _findings_index_sheet(summary: dict, detail_sheet_names: set[str]) -> pd.DataFrame:
    """Master index: one row per finding, with detail-sheet cross-reference."""
    rows = []
    for finding in _ordered_findings(summary):
        check_key  = str(finding.get("check_key", ""))
        check_name = str(finding.get("check_name", ""))
        sev        = str(finding.get("severity", "")).upper()
        field      = str(finding.get("field", ""))
        ra         = _records_affected(finding)

        issue_name, _, _, detail_sheet = _meta(check_key, check_name, field)

        rows.append({
            "Issue Name":       issue_name,
            "Severity":         sev,
            "Records Affected": ra,
            "Detail Sheet":     detail_sheet if detail_sheet in detail_sheet_names else "-",
            "Description":      str(finding.get("description", "")),
        })

    if not rows:
        return pd.DataFrame([{"Note": "No findings recorded."}])
    return pd.DataFrame(rows)


# ── Sheet 6: Data_Quality_Score ───────────────────────────────────────────────

def _data_quality_score_sheet(summary: dict) -> pd.DataFrame:
    sev_counts = summary.get("severity_counts") or {}
    rows = [
        {"Metric": "Total Records",           "Value": summary.get("total_rows", 0)},
        {"Metric": "Completeness %",          "Value": summary.get("overall_completeness", 0)},
        {"Metric": "Gate Status",             "Value": summary.get("gate_status", "-")},
        {"Metric": "Gate Message",            "Value": summary.get("gate_message", "")},
        {"Metric": "Critical Issues",         "Value": sev_counts.get("CRITICAL", 0)},
        {"Metric": "High Issues",             "Value": sev_counts.get("HIGH", 0)},
        {"Metric": "Medium Issues",           "Value": sev_counts.get("MEDIUM", 0)},
        {"Metric": "Low Issues",              "Value": sev_counts.get("LOW", 0)},
        {"Metric": "Source File",             "Value": summary.get("source_filename", "")},
    ]
    return pd.DataFrame(rows)


# ── Sheet assembly ────────────────────────────────────────────────────────────

def _build_sheets(
    file_path: Path, sheet_name: int | str, run_dir: Path
) -> dict[str, pd.DataFrame]:
    summary = _load_summary(run_dir)
    df, _, row_annotations = _read_and_normalize(file_path, sheet_name=sheet_name)

    # Build detail sheets first (to know which sheet names exist for Findings_Index)
    detail_sheets = _build_findings_detail_sheets(df, summary, row_annotations, run_dir)

    sheets: dict[str, pd.DataFrame] = {}

    # 1. Executive_Summary
    sheets["Executive_Summary"] = _fmt(_executive_summary_sheet(summary))

    # 2. Fix_List
    sheets["Fix_List"] = _fmt(_fix_list_sheet(summary))

    # 3. Fix_List_Detail
    sheets["Fix_List_Detail"] = _fix_list_detail_sheet(df, summary, row_annotations)

    # 4. Per-issue Findings_* detail tabs (sorted by PDF order)
    sheets.update(detail_sheets)

    # 5. Findings_Index
    sheets["Findings_Index"] = _fmt(_findings_index_sheet(summary, set(detail_sheets)))

    # 6. Data_Quality_Score
    sheets["Data_Quality_Score"] = _fmt(_data_quality_score_sheet(summary))

    return sheets


# ── Excel styling pass ────────────────────────────────────────────────────────

def _sev_fill(sev: str) -> PatternFill | None:
    color = SEV_FILLS.get(str(sev).upper())
    if not color:
        return None
    return PatternFill(fill_type="solid", fgColor=color)


def _autosize_and_style(out_path: Path, sev_col_sheets: dict[str, list[str]]) -> None:
    """
    sev_col_sheets: maps sheet_name → list of column headers that contain severity values
                    (used to color-code those cells).
    """
    wb = load_workbook(out_path)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        # Header row
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Column auto-width (sample first 200 rows for speed)
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            length = 0
            for cell in col_cells[:200]:
                try:
                    length = max(length, len(str(cell.value or "")))
                except Exception:
                    continue
            ws.column_dimensions[letter].width = min(max(length + 2, 12), 50)

        # Severity-based row / cell coloring
        sev_cols = sev_col_sheets.get(ws.title, [])
        if not sev_cols:
            continue

        # Build column-index map from headers
        col_idx: dict[str, int] = {}   # header → 0-based index
        for cell in ws[1]:
            h = str(cell.value or "").strip()
            if h:
                col_idx[h] = cell.column - 1

        for row_cells in ws.iter_rows(min_row=2):
            sev_val = ""
            for sev_col in sev_cols:
                idx = col_idx.get(sev_col)
                if idx is not None and idx < len(row_cells):
                    sev_val = str(row_cells[idx].value or "").strip().upper()
                    if sev_val:
                        break

            fill = _sev_fill(sev_val)
            if fill:
                for cell in row_cells:
                    cell.fill = fill

    wb.save(out_path)


# ── Entry points ──────────────────────────────────────────────────────────────

def build_workbook(
    file_path: Path, run_dir: Path, out_path: Path, sheet_name: int | str = 0
) -> None:
    sheets = _build_sheets(file_path, sheet_name, run_dir)

    # Sheets where severity cells should drive row coloring
    sev_col_sheets = {
        "Executive_Summary": ["Severity"],
        "Fix_List":          ["Severity"],
        "Findings_Index":    ["Severity"],
    }

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for key, frame in sheets.items():
            frame.to_excel(writer, sheet_name=key, index=False)

    _autosize_and_style(out_path, sev_col_sheets)
    print(f"[build_internal_audit_workbook] wrote: {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build internal audit workbook")
    parser.add_argument("--file",       required=True, help="Source file used for the audit")
    parser.add_argument("--run-dir",    required=True, help="Directory containing internal audit outputs")
    parser.add_argument("--out",        required=True, help="Output workbook path")
    parser.add_argument("--sheet-name", default="0",  help="Excel sheet index or name")
    args = parser.parse_args()
    sheet_name: int | str = (
        int(args.sheet_name) if str(args.sheet_name).lstrip("-").isdigit() else args.sheet_name
    )
    build_workbook(Path(args.file), Path(args.run_dir), Path(args.out), sheet_name=sheet_name)


if __name__ == "__main__":
    main()

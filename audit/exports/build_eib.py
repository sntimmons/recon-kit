"""
build_eib.py - Build Workday EIB Excel files from approved corrections.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from connectors.workday import Workday

TODAY = date.today().isoformat()


def _instructions(ws, run_id: str, correction_type: str, count: int) -> None:
    lines = [
        "Data Whisperer - Workday EIB Export",
        f"Run ID: {run_id}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "Target system: Workday",
        f"Correction type: {correction_type}",
        f"Record count: {count}",
        "",
        "IMPORTANT - Review all records before loading.",
        "This file requires CHRO authorization before use.",
        f"See the CHRO Approval Document for run {run_id}.",
        "",
        "Do not modify column headers in the Data sheet.",
        "Workday will reject the file if headers are changed.",
    ]
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)


def _write_workbook(path: Path, run_id: str, correction_type: str, headers: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws_i = wb.active
    ws_i.title = "Instructions"
    _instructions(ws_i, run_id, correction_type, len(rows))
    ws_d = wb.create_sheet("Data")
    ws_d.append(headers)
    for row in rows:
        ws_d.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _job_org_rows(df: pd.DataFrame) -> list[list]:
    rows = []
    for _, row in df.iterrows():
        rows.append([
            row.get("worker_id", ""),
            row.get("effective_date", TODAY),
            row.get("position", ""),
            row.get("position", ""),
            row.get("district", ""),
            "",
            row.get("location", "") or row.get("location_state", ""),
        ])
    return rows


def _hire_date_rows(df: pd.DataFrame) -> list[list]:
    rows = []
    for _, row in df.iterrows():
        hire_date = row.get("hire_date", "")
        rows.append([
            row.get("worker_id", ""),
            hire_date,
            hire_date,
        ])
    return rows


def _status_rows(df: pd.DataFrame) -> list[list]:
    rows = []
    for _, row in df.iterrows():
        rows.append([
            row.get("worker_id", ""),
            row.get("effective_date", TODAY),
            row.get("worker_status", ""),
            "",
            "",
            "",
        ])
    return rows


def _salary_rows(df: pd.DataFrame) -> list[list]:
    rows = []
    for _, row in df.iterrows():
        rows.append([
            row.get("worker_id", ""),
            row.get("effective_date", TODAY),
            row.get("compensation_amount", ""),
            row.get("currency", "USD") or "USD",
            "Salary",
            "Annual",
        ])
    return rows


def build_eibs(run_id: str, run_dir: Path) -> list[Path]:
    workday = Workday()
    created: list[Path] = []
    file_specs = [
        ("job_org", "corrections_job_org.csv", "eib_job_org.xlsx", workday.eib_job_org_columns, _job_org_rows),
        ("hire_date", "corrections_hire_date.csv", "eib_hire_date.xlsx", workday.eib_hire_date_columns, _hire_date_rows),
        ("status", "corrections_status.csv", "eib_status.xlsx", workday.eib_status_columns, _status_rows),
        ("salary", "corrections_salary.csv", "eib_salary.xlsx", workday.eib_salary_columns, _salary_rows),
    ]

    for correction_type, src_name, out_name, headers, builder in file_specs:
        src = run_dir / src_name
        if not src.exists():
            continue
        df = pd.read_csv(src)
        if df.empty:
            continue
        out_path = run_dir / out_name
        _write_workbook(out_path, run_id, correction_type, headers, builder(df))
        created.append(out_path)
    return created


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Build Workday EIB exports from correction CSVs.")
    parser.add_argument("--run-id", required=True, help="Run identifier")
    parser.add_argument("--run-dir", required=True, help="Run output directory")
    args = parser.parse_args(argv)

    created = build_eibs(args.run_id, Path(args.run_dir))
    for path in created:
        print(f"[build_eib] wrote: {path}")


if __name__ == "__main__":
    main()

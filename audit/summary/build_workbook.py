"""
build_workbook.py - Full Excel workbook export.

Sheets
------
  1) Summary              - key statistics
  2) All_Matches          - full dataset (all rows, all wide_compare columns)
  3) Salary_Mismatches    - rows where fix_types contains "salary"
  4) Status_Mismatches    - rows where fix_types contains "status"
  5) HireDate_Mismatches  - rows where fix_types contains "hire_date"
  6) JobOrg_Mismatches    - rows where fix_types contains "job_org"
  7) Review_Queue         - review_queue.csv if present, else action==REVIEW filter
  8) Corrections_Manifest - corrections_manifest.csv if present, else placeholder

Source priority for All_Matches data:
  1. audit/exports/out/wide_compare.csv  (preferred)
  2. matched_pairs view in audit.db      (fallback, computes gating on the fly)

Uses openpyxl write_only=True mode for streaming writes - no MemoryError on large
datasets.  Formatting: bold/coloured header row via WriteOnlyCell; freeze_panes and
auto_filter are not available in write_only mode.

Run:
    venv/Scripts/python.exe audit/summary/build_workbook.py [--out PATH] [--wide PATH] [--db PATH]
"""
from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.cell.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "[error] openpyxl not installed. "
        "Run: venv/Scripts/pip.exe install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)

_HERE = Path(__file__).resolve().parent    # audit/summary/
sys.path.insert(0, str(_HERE))

from gating import (
    classify_all,
    salary_delta,
    payrate_delta,
    build_summary_str,
    _parse_confidence,
    _norm,
)
from config_loader import load_policy, load_pii_config
from explanation import generate_explanation
from sanity_checks import detect_wave_dates

import os as _os_rk
ROOT         = _HERE.parents[1]
_rk_work     = Path(_os_rk.environ["RK_WORK_DIR"]) if "RK_WORK_DIR" in _os_rk.environ else None
DB_PATH      = ROOT / "audit" / "audit.db"
WIDE_CSV     = ROOT / "audit" / "exports" / "out" / "wide_compare.csv"
REVIEW_CSV   = (_rk_work / "review_queue.csv")                                                  if _rk_work else (_HERE / "review_queue.csv")
MANIFEST_CSV = (_rk_work / "audit" / "corrections" / "out" / "corrections_manifest.csv")       if _rk_work else (ROOT / "audit" / "corrections" / "out" / "corrections_manifest.csv")
HELD_CSV     = (_rk_work / "audit" / "corrections" / "out" / "held_corrections.csv")           if _rk_work else (ROOT / "audit" / "corrections" / "out" / "held_corrections.csv")
OUT_PATH     = _HERE / "recon_workbook.xlsx"

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
_HDR_FILL      = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
_HDR_FONT      = Font(bold=True, color="FFFFFF")
_ALT_FILL      = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")
_SUM_HDR_FONT  = Font(bold=True, size=11)
# Red header for critical warning sheets (Active/$0 salary)
_CRIT_HDR_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
_CRIT_HDR_FONT = Font(bold=True, color="FFFFFF")
# Orange header for rejected-match sheet
_REJ_HDR_FILL  = PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid")
_REJ_HDR_FONT  = Font(bold=True, color="FFFFFF")
# Amber header for held-corrections sheet
_HELD_HDR_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
_HELD_HDR_FONT = Font(bold=True, color="000000")
# Light-gray header for reference-only sheets (Clean_Data, Unmatched)
_REF_HDR_FILL  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_REF_HDR_FONT  = Font(bold=True, color="000000")

# ---------------------------------------------------------------------------
# Column label mapping (snake_case -> human-readable Title Case)
# ---------------------------------------------------------------------------
_COL_LABELS: dict[str, str] = {
    "pair_id":             "Pair ID",
    "match_source":        "Match Source",
    "confidence":          "Confidence",
    "action":              "Action",
    "reason":              "Reason",
    "fix_types":           "Fix Types",
    "summary":             "Summary",
    "match_explanation":   "Match Explanation",
    "priority_score":      "Priority Score",
    "priority_reason":     "Priority Reason",
    "old_worker_id":       "Worker ID (Old)",
    "new_worker_id":       "Worker ID (New)",
    "old_first_name_norm": "First Name",
    "new_first_name_norm": "New First Name",
    "old_last_name_norm":  "Last Name",
    "new_last_name_norm":  "New Last Name",
    "old_full_name_norm":  "Full Name (Old)",
    "new_full_name_norm":  "Full Name (New)",
    "old_worker_status":   "Status (Old)",
    "new_worker_status":   "Status (New)",
    "old_worker_type":     "Worker Type (Old)",
    "new_worker_type":     "Worker Type (New)",
    "old_salary":          "Salary (Old)",
    "new_salary":          "Salary (New)",
    "salary_delta":        "Salary Change",
    "salary_ratio":        "Salary Ratio",
    "old_payrate":         "Pay Rate (Old)",
    "new_payrate":         "Pay Rate (New)",
    "payrate_delta":       "Pay Rate Change",
    "old_hire_date":       "Hire Date (Old)",
    "new_hire_date":       "Hire Date (New)",
    "old_position":        "Position (Old)",
    "new_position":        "Position (New)",
    "old_district":        "District (Old)",
    "new_district":        "District (New)",
    "old_location_state":  "State (Old)",
    "new_location_state":  "State (New)",
    "old_location":        "Location (Old)",
    "new_location":        "Location (New)",
    "hire_date_pattern":   "Hire Date Pattern",
    "status_changed":      "Status Changed",
    "hire_date_changed":   "Hire Date Changed",
    "job_org_changed":     "Job/Org Changed",
    "needs_review":        "Needs Review",
    "suggested_action":    "Suggested Action",
    # held corrections
    "worker_id":           "Worker ID",
    "overall_action":      "Action",
    "hold_reason":         "Hold Reason",
    # corrections manifest
    "correction_id":       "Correction ID",
    "correction_type":     "Correction Type",
    "field":               "Field",
    "old_value":           "Old Value",
    "new_value":           "New Value",
    "fix_description":     "Fix Description",
    "recommended_action":  "Recommended Action",
}

_ACTION_LABELS: dict[str, str] = {
    "APPROVE":      "Safe",
    "REVIEW":       "Needs Review",
    "REJECT_MATCH": "Wrong Match",
}


def _label(col: str) -> str:
    """Return human-readable label for a column name."""
    return _COL_LABELS.get(col, col.replace("_", " ").title())


def _transform_action_val(val) -> str:
    """Map APPROVE/REVIEW/REJECT_MATCH to Safe/Needs Review/Wrong Match."""
    if val is None:
        return ""
    s = str(val).strip().upper()
    return _ACTION_LABELS.get(s, str(val))


def _transform_fix_types_val(val) -> str:
    """Convert 'salary|status' -> 'Salary, Status'. Empty -> 'No Changes'."""
    if not val or str(val).strip() in ("", "nan"):
        return "No Changes"
    parts = [p.strip() for p in str(val).split("|") if p.strip()]
    _ft_labels = {
        "salary":    "Salary",
        "status":    "Status",
        "hire_date": "Hire Date",
        "job_org":   "Job / Org",
    }
    return ", ".join(_ft_labels.get(p, p.replace("_", " ").title()) for p in parts)


def _slim(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    """Return df restricted to requested columns that exist, in order."""
    keep = [c for c in cols if c in df.columns]
    return df[keep].copy()


def _apply_labels(df: pd.DataFrame) -> pd.DataFrame:
    """Transform action/fix_types values and rename columns to human-readable labels."""
    result = df.copy()
    if "action" in result.columns:
        result["action"] = result["action"].apply(_transform_action_val)
    if "overall_action" in result.columns:
        result["overall_action"] = result["overall_action"].apply(_transform_action_val)
    if "fix_types" in result.columns:
        result["fix_types"] = result["fix_types"].apply(_transform_fix_types_val)
    result.columns = [_label(c) for c in result.columns]
    return result


def _lookup_names_from_wide(worker_ids, all_df: pd.DataFrame) -> dict:
    """Build {str(worker_id): (first_name, last_name)} lookup from all_df. O(n)."""
    result: dict = {}
    if all_df.empty or not worker_ids:
        return result
    id_set = {str(w) for w in worker_ids if w}
    id_col = "old_worker_id" if "old_worker_id" in all_df.columns else None
    fn_col = "old_first_name_norm" if "old_first_name_norm" in all_df.columns else None
    ln_col = "old_last_name_norm" if "old_last_name_norm" in all_df.columns else None
    if not id_col:
        return result
    for row in all_df[[c for c in [id_col, fn_col, ln_col] if c]].itertuples(index=False, name=None):
        idx = 0
        wid = str(row[idx])
        if wid in id_set and wid not in result:
            first = str(row[1]) if fn_col and len(row) > 1 else ""
            last  = str(row[2]) if ln_col and len(row) > 2 else ""
            result[wid] = (first if first != "nan" else "", last if last != "nan" else "")
    return result


# ---------------------------------------------------------------------------
# Slim column lists for each mismatch category
# ---------------------------------------------------------------------------
_SALARY_SLIM_COLS = [
    "pair_id", "match_source", "confidence", "action", "priority_score",
    "old_worker_id", "old_first_name_norm", "old_last_name_norm",
    "old_worker_status", "new_worker_status",
    "old_salary", "new_salary", "salary_delta",
    "old_payrate", "new_payrate", "payrate_delta",
    "fix_types", "reason", "summary",
]

_STATUS_SLIM_COLS = [
    "pair_id", "match_source", "confidence", "action", "priority_score",
    "old_worker_id", "old_first_name_norm", "old_last_name_norm",
    "old_worker_status", "new_worker_status",
    "old_worker_type", "new_worker_type",
    "fix_types", "reason", "summary",
]

_HIRE_DATE_SLIM_COLS = [
    "pair_id", "match_source", "confidence", "action", "priority_score",
    "old_worker_id", "old_first_name_norm", "old_last_name_norm",
    "old_worker_status", "new_worker_status",
    "old_hire_date", "new_hire_date",
    "hire_date_pattern",
    "fix_types", "reason", "summary",
]

_JOB_ORG_SLIM_COLS = [
    "pair_id", "match_source", "confidence", "action", "priority_score",
    "old_worker_id", "old_first_name_norm", "old_last_name_norm",
    "old_worker_status", "new_worker_status",
    "old_position", "new_position",
    "old_district", "new_district",
    "old_location_state", "new_location_state",
    "fix_types", "reason", "summary",
]


# ---------------------------------------------------------------------------
# Sheet writer helpers (standard workbook with styling)
# ---------------------------------------------------------------------------

def _style_header(cells, font=None, fill=None) -> None:
    for cell in cells:
        cell.font = font or _HDR_FONT
        cell.fill = fill or _HDR_FILL


def _write_df_to_sheet(ws, df: pd.DataFrame, hdr_font=None, hdr_fill=None) -> None:
    """Write DataFrame with styled header, filters, freeze panes, and banded rows."""
    cols = list(df.columns)
    if not cols:
        return

    ws.append(cols)
    _style_header(ws[1], font=hdr_font or _HDR_FONT, fill=hdr_fill or _HDR_FILL)

    for idx, row_tuple in enumerate(df.itertuples(index=False, name=None), start=2):
        ws.append(list(row_tuple))
        if idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = _ALT_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_df_to_sheet_styled(ws, df: pd.DataFrame, hdr_font=None, hdr_fill=None) -> None:
    _write_df_to_sheet(ws, df, hdr_font=hdr_font, hdr_fill=hdr_fill)


def _write_mismatch_slim(ws, df: pd.DataFrame, slim_cols: list[str]) -> None:
    """Write a mismatch sheet slimmed to relevant columns with human-readable labels."""
    out = _apply_labels(_slim(df, slim_cols))
    _write_df_to_sheet(ws, out)


def _write_held_corrections_sheet(ws, held_df: pd.DataFrame, all_df: pd.DataFrame) -> None:
    """Write Held_Corrections sheet with amber header and name lookup from all_df."""
    if held_df.empty:
        ws.append(["No held corrections found."])
        return
    df = held_df.copy()
    # Inject First Name / Last Name from all_df lookup
    if "worker_id" in df.columns:
        wids = df["worker_id"].astype(str).tolist()
        name_map = _lookup_names_from_wide(wids, all_df)
        insert_at = 1
        df.insert(insert_at,     "first_name", df["worker_id"].astype(str).map(lambda w: name_map.get(w, ("", ""))[0]))
        df.insert(insert_at + 1, "last_name",  df["worker_id"].astype(str).map(lambda w: name_map.get(w, ("", ""))[1]))
    df = _apply_labels(df)
    cols = list(df.columns)
    ws.append(cols)
    _style_header(ws[1], font=_HELD_HDR_FONT, fill=_HELD_HDR_FILL)
    for idx, row_tuple in enumerate(df.itertuples(index=False, name=None), start=2):
        ws.append(list(row_tuple))
        if idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = _ALT_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_review_queue_slim(ws, review_df: pd.DataFrame, all_df: pd.DataFrame) -> None:
    """Write Review_Queue sheet slimmed to 8 key columns with derived Recommended Action."""
    if review_df.empty:
        ws.append([WriteOnlyCell(ws, value="No records in review queue.")])
        return
    df = review_df.copy()
    # Merge missing columns from all_df via pair_id
    if "pair_id" in df.columns and not all_df.empty:
        want = [c for c in ["pair_id", "action", "fix_types", "summary", "reason"]
                if c in all_df.columns and c not in df.columns]
        if want:
            df = df.merge(all_df[["pair_id"] + [c for c in want if c != "pair_id"]],
                          on="pair_id", how="left")
    # Derive Recommended Action
    def _recommended(row) -> str:
        act = str(row.get("action", "")).strip().upper()
        ft  = str(row.get("fix_types", "")).strip().lower()
        if act == "REJECT_MATCH":
            return "Verify worker identity - match may be wrong"
        if act == "APPROVE":
            return "No action needed - auto-approved"
        fixes = [p.strip() for p in ft.split("|") if p.strip()]
        parts = []
        if "salary"    in fixes: parts.append("review salary change")
        if "status"    in fixes: parts.append("confirm status")
        if "hire_date" in fixes: parts.append("verify hire date")
        if "job_org"   in fixes: parts.append("check position / org")
        return "Manually review: " + ", ".join(parts) if parts else "Manually review"

    df["recommended_action"] = df.apply(_recommended, axis=1)
    slim_cols = [
        "old_worker_id", "old_first_name_norm", "old_last_name_norm",
        "action", "fix_types", "summary", "recommended_action", "priority_score",
    ]
    df = _apply_labels(_slim(df, slim_cols))
    _write_df_to_sheet(ws, df)


def _write_unmatched_sheet(ws, unmatched_df: pd.DataFrame) -> None:
    """Write an unmatched-records sheet with light-gray reference header."""
    if unmatched_df.empty:
        ws.append(["No unmatched records found."])
        return
    _write_df_to_sheet_styled(ws, unmatched_df, hdr_font=_REF_HDR_FONT, hdr_fill=_REF_HDR_FILL)


def _enhance_manifest(manifest_df: pd.DataFrame) -> pd.DataFrame:
    """Add Fix Description column and transform action/fix_types to human-readable values."""
    if manifest_df.empty or "note" in manifest_df.columns:
        return manifest_df  # placeholder or blocked row, leave as-is
    df = manifest_df.copy()
    # Transform action columns
    for col in ("action", "overall_action"):
        if col in df.columns:
            df[col] = df[col].apply(_transform_action_val)
    # Transform fix_types
    if "fix_types" in df.columns:
        df["fix_types"] = df["fix_types"].apply(_transform_fix_types_val)
    # Add fix_description if absent
    if "fix_description" not in df.columns:
        old_col = next((c for c in ("old_value", "from_value", "old_salary", "old_worker_status",
                                     "old_hire_date", "old_position") if c in df.columns), None)
        new_col = next((c for c in ("new_value", "to_value", "new_salary", "new_worker_status",
                                     "new_hire_date", "new_position") if c in df.columns), None)
        ct_col  = next((c for c in ("correction_type", "fix_type", "fix_types") if c in df.columns), None)
        _ct_map = {
            "salary":    "Update salary",
            "status":    "Change status",
            "hire_date": "Update hire date",
            "job_org":   "Update position / org",
        }
        def _desc(row) -> str:
            ct  = str(row.get(ct_col, "")).strip().lower() if ct_col else ""
            old = row.get(old_col, "") if old_col else ""
            new = row.get(new_col, "") if new_col else ""
            base = _ct_map.get(ct, "Update field")
            if old and new:
                return f"{base}: {old} -> {new}"
            return base
        df["fix_description"] = df.apply(_desc, axis=1)
    return df


def validate_active_zero_salary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return rows where new_worker_status is 'active' and new_salary is $0 or blank.

    These records indicate a mapping failure or data artefact - staging a salary
    correction for them would zero out an active employee's pay in the target system.
    """
    if df.empty:
        return df.iloc[0:0]  # empty with same columns
    mask_status = (
        df.get("new_worker_status", pd.Series(dtype=str))
        .fillna("")
        .str.strip()
        .str.lower()
        .isin(["active", ""])
    )
    new_sal_num = pd.to_numeric(
        df.get("new_salary", pd.Series(dtype=object))
        .astype(str)
        .str.replace(",", "")
        .str.replace("$", ""),
        errors="coerce",
    )
    mask_zero = new_sal_num.isna() | (new_sal_num == 0.0)
    return df[mask_status & mask_zero].copy()


def _load_salary_parse_stats() -> dict[str, object]:
    """Read salary parse failure counts/samples from mapping reports if present."""
    _run_outs = (_rk_work / "outputs") if _rk_work else (ROOT / "outputs")

    def _one(side: str) -> tuple[int, list[str]]:
        p = _run_outs / f"mapping_report_mapped_{side}.json"
        if not p.exists():
            return 0, []
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            count = int(data.get("salary_parse_failures", 0) or 0)
            samples = [str(x) for x in (data.get("salary_parse_failure_samples", []) or [])]
            return count, samples
        except Exception:
            return 0, []

    old_count, old_samples = _one("old")
    new_count, new_samples = _one("new")
    return {
        "old_count": old_count,
        "new_count": new_count,
        "old_samples": old_samples,
        "new_samples": new_samples,
    }


def _fix_types_text(df: pd.DataFrame) -> pd.Series | None:
    if "fix_types" not in df.columns:
        return None
    return df["fix_types"].fillna("").astype("string")


def _write_summary_sheet(
    ws,
    all_df: pd.DataFrame,
    db_path: Path,
    wide_src: str,
    unmatched_old: int = 0,
    unmatched_new: int = 0,
    salary_parse_stats: dict[str, object] | None = None,
) -> None:
    """Write the Summary sheet as a key-value table (write_only compatible)."""

    def _kv(label: str = "", value="", bold: bool = False, font=None) -> None:
        if (bold or font) and label:
            lc = WriteOnlyCell(ws, value=label)
            lc.font = font or _HDR_FONT
            ws.append([lc, value])
        else:
            ws.append([label, value])

    # ------------------------------------------------------------------
    # START HERE section - 5-step guide + priority score legend
    # ------------------------------------------------------------------
    def _bold_cell(text: str) -> WriteOnlyCell:
        c = WriteOnlyCell(ws, value=text)
        c.font = Font(bold=True, size=12)
        return c

    def _nav_cell(text: str) -> WriteOnlyCell:
        c = WriteOnlyCell(ws, value=text)
        c.font = Font(bold=True, color="0070C0")
        return c

    ws.append([_bold_cell("START HERE - How to use this workbook")])
    ws.append([])
    ws.append(["Step 1", _nav_cell("CRITICAL_Zero_Salary"),
               "Active workers with $0 salary - fix before any import"])
    ws.append(["Step 2", _nav_cell("Held_Corrections"),
               "Records held back: Wrong Match or Needs Review - requires manual decision"])
    ws.append(["Step 3", _nav_cell("Review_Queue"),
               "Prioritized list of records needing attention - work top to bottom"])
    ws.append(["Step 4", _nav_cell("Salary / Status / HireDate / JobOrg sheets"),
               "Mismatch detail by category - use for context and spot-checking"])
    ws.append(["Step 5", _nav_cell("Corrections_Manifest"),
               "Apply Safe corrections to target system"])
    ws.append([])
    ws.append([_bold_cell("Priority Score Legend")])
    ws.append(["70+",   "Critical - must review before import"])
    ws.append(["40-69", "High - review recommended"])
    ws.append(["20-39", "Medium - spot check advised"])
    ws.append(["0-19",  "Low - auto-approve candidate"])
    ws.append([])
    ws.append([_bold_cell("Action Labels")])
    ws.append(["Safe",         "No issues found - approved for import"])
    ws.append(["Needs Review", "Flagged for manual review before import"])
    ws.append(["Wrong Match",  "Match confidence too low - verify identity"])
    ws.append([])

    # ------------------------------------------------------------------
    # Existing summary statistics
    # ------------------------------------------------------------------
    _kv("Reconciliation Workbook", "", font=_SUM_HDR_FONT)
    _kv("Generated", datetime.now().strftime("%Y-%m-%d %H:%M"))
    _kv("Source DB", str(db_path.name))
    _kv("Data source", wide_src)
    _kv()

    total = len(all_df)
    fix_types_text = _fix_types_text(all_df)
    _kv("MATCHED PAIRS", "", bold=True)
    _kv("  Total rows", total)
    if "action" in all_df.columns:
        n_approve      = int((all_df["action"] == "APPROVE").sum())
        n_review       = int((all_df["action"] == "REVIEW").sum())
        n_reject_match = int((all_df["action"] == "REJECT_MATCH").sum())
        _kv("  APPROVE", n_approve)
        _kv("  REVIEW",  n_review)
        if n_reject_match:
            _kv("  REJECT_MATCH", n_reject_match)
    _kv()

    _kv("UNMATCHED RECORDS", "", bold=True)
    _kv("  Unmatched old system records", unmatched_old)
    _kv("  Unmatched new system records", unmatched_new)
    _kv()

    salary_parse_stats = salary_parse_stats or {}
    _kv("DATA PARSE CHECKS", "", bold=True)
    _kv("  Salary parse failures (old system)", int(salary_parse_stats.get("old_count", 0) or 0))
    _kv("  Salary parse failures (new system)", int(salary_parse_stats.get("new_count", 0) or 0))
    _kv()

    if "match_source" in all_df.columns:
        _kv("BY MATCH SOURCE", "", bold=True)
        for src, cnt in all_df["match_source"].value_counts().items():
            _kv(f"  {src}", int(cnt))
        # Fuzzy matches (sub-1.0 confidence)
        if "confidence" in all_df.columns:
            conf_num = pd.to_numeric(all_df["confidence"], errors="coerce")
            n_fuzzy  = int((conf_num < 1.0).sum())
            _kv("  Fuzzy (confidence < 1.0)", n_fuzzy)
        _kv()

    if fix_types_text is not None:
        _kv("MISMATCH TYPES", "", bold=True)
        for ft, label in [
            ("salary",    "Salary changes"),
            ("status",    "Status changes"),
            ("hire_date", "Hire-date changes"),
            ("job_org",   "Job/org changes"),
        ]:
            cnt = int(fix_types_text.str.contains(ft, na=False).sum())
            _kv(f"  {label}", cnt)
        no_change = int((fix_types_text == "").sum())
        _kv("  No changes", no_change)
        _kv()

    if "salary_delta" in all_df.columns:
        # Fix 5: count rows where fix_types contains "salary" (not just non-null salary_delta)
        if fix_types_text is not None:
            sal_rows_count = int(fix_types_text.str.contains("salary", na=False).sum())
        else:
            sal_rows_count = None

        # Active/$0 records are data quality issues (missing/bad data from source),
        # not real salary changes.  Their delta is -(old_salary), e.g. -$50,000 for
        # a worker whose new file has $0 - including them collapses mean/median and
        # masks the true distribution of legitimate salary corrections.
        # Identify them once, use for both exclusion and the CRITICAL warning below.
        active_zero_df = validate_active_zero_salary(all_df)
        n_az = len(active_zero_df)

        sal_d = pd.to_numeric(all_df["salary_delta"], errors="coerce")
        # Drop Active/$0 records from the stats series by index alignment
        if n_az > 0:
            sal_d = sal_d.drop(active_zero_df.index, errors="ignore")

        sal_d_nonzero = sal_d[sal_d != 0].dropna()
        if len(sal_d_nonzero) > 0 or sal_rows_count:
            _kv("SALARY DELTA STATS", "", bold=True)
            if n_az > 0:
                _kv(f"  (excl. {n_az} Active/$0 - data quality, not real changes)", "")
            if sal_rows_count is not None:
                _kv("  Rows with salary change", sal_rows_count)
            if len(sal_d_nonzero) > 0:
                _kv("  Mean delta",   round(float(sal_d_nonzero.mean()), 2))
                _kv("  Median delta", round(float(sal_d_nonzero.median()), 2))
                _kv("  Max increase", round(float(sal_d_nonzero.max()), 2))
                _kv("  Max decrease", round(float(sal_d_nonzero.min()), 2))
            # Active/$0 salary warning
            if n_az > 0:
                lc = WriteOnlyCell(ws, value=f"  CRITICAL: Active workers with $0 salary")
                lc.font = Font(bold=True, color="C00000")
                ws.append([lc, n_az])
            _kv()

    _kv("SHEETS IN THIS WORKBOOK", "", bold=True)
    _kv("  Review_Queue",         "Prioritized records requiring manual decision")
    _kv("  Held_Corrections",     "Blocked / wrong-match records - decide before import")
    _kv("  CRITICAL_Zero_Salary", "Active workers with $0 or missing salary")
    if "fix_types" in all_df.columns:
        for sheet, ft in [
            ("  Salary_Mismatches",   "salary"),
            ("  Status_Mismatches",   "status"),
            ("  HireDate_Mismatches", "hire_date"),
            ("  JobOrg_Mismatches",   "job_org"),
        ]:
            cnt = int(all_df["fix_types"].str.contains(ft, na=False).sum())
            _kv(sheet, cnt)
    _kv("  Corrections_Manifest",  "Auto-approved corrections ready to apply")
    _kv("  Unmatched_Old",         unmatched_old)
    _kv("  Unmatched_New",         unmatched_new)
    _kv("  Clean_Data",            f"{total:,} rows - full dataset reference")
    _kv("  All_Matches",           f"{total:,} rows - full matched pairs with all columns")


# ---------------------------------------------------------------------------
# Fallback: compute wide_compare from DB
# ---------------------------------------------------------------------------

def _str_eq(a, b) -> bool:
    return _norm(a) == _norm(b)


def _salary_ratio(old_sal, new_sal):
    try:
        o = float(str(old_sal or "").replace(",", "").replace("$", ""))
        n = float(str(new_sal or "").replace(",", "").replace("$", ""))
        return None if o == 0 else round(n / o, 6)
    except Exception:
        return None


def _priority_score(row: dict, fix_types: list[str], sal_d, result: dict) -> int:
    score = 0
    if "status" in fix_types:
        score += 50
    if sal_d is not None:
        if abs(sal_d) >= 5000:
            score += 30
        if abs(sal_d) >= 1000:
            score += 15
    if "hire_date" in fix_types:
        score += 20
    if "job_org" in fix_types:
        if not _str_eq(row.get("old_position"), row.get("new_position")):
            score += 10
        if not _str_eq(row.get("old_district"), row.get("new_district")):
            score += 8
        if not _str_eq(row.get("old_location_state"), row.get("new_location_state")):
            score += 6
    if _norm(row.get("match_source", "")) != "worker_id":
        score += 10
    if _parse_confidence(row.get("confidence")) is None:
        score += 10
    if len(fix_types) > 1:
        score += 5
    return score


def _load_wide_from_db(db_path: Path) -> pd.DataFrame:
    """Compute wide_compare columns directly from matched_pairs (fallback)."""
    print("[build_workbook] computing gating from DB (this may take a moment) ...")
    con = sqlite3.connect(str(db_path))
    try:
        mp = pd.read_sql_query("SELECT * FROM matched_pairs", con)
    finally:
        con.close()

    if "confidence" not in mp.columns:
        mp["confidence"] = None

    # PII guard: suppress DOB from DB-sourced wide data if configured.
    pii_cfg = load_pii_config(load_policy())
    if not pii_cfg.get("include_dob_in_exports", True):
        dob_cols = [c for c in ["old_dob", "new_dob"] if c in mp.columns]
        if dob_cols:
            mp = mp.drop(columns=dob_cols)
            print(f"[build_workbook] [pii] suppressed DOB columns: {dob_cols}")

    has_loc = "new_location" in mp.columns
    has_wt  = "old_worker_type" in mp.columns

    all_rows   = mp.to_dict(orient="records")
    wave_dates = detect_wave_dates(all_rows)

    rows = []
    for r in all_rows:
        result    = classify_all(r, wave_dates=wave_dates)
        fix_types = result["fix_types"]
        sal_d     = salary_delta(r)
        pay_d     = payrate_delta(r)
        prio      = _priority_score(r, fix_types, sal_d, result)

        rows.append({
            "pair_id":            r.get("pair_id", ""),
            "match_source":       r.get("match_source", ""),
            "confidence":         r.get("confidence"),
            "action":             result["action"],
            "reason":             result["reason"],
            "fix_types":          "|".join(fix_types),
            "summary":            build_summary_str(r, fix_types) if fix_types else "no_changes",
            "match_explanation":  generate_explanation(r, result),
            "priority_score":     prio,
            "old_full_name_norm": r.get("old_full_name_norm", ""),
            "new_full_name_norm": r.get("new_full_name_norm", ""),
            "old_first_name_norm": r.get("old_first_name_norm", ""),
            "new_first_name_norm": r.get("new_first_name_norm", ""),
            "old_last_name_norm":  r.get("old_last_name_norm", ""),
            "new_last_name_norm":  r.get("new_last_name_norm", ""),
            "old_middle_name":     r.get("old_middle_name", ""),
            "new_middle_name":     r.get("new_middle_name", ""),
            "old_suffix":          r.get("old_suffix", ""),
            "new_suffix":          r.get("new_suffix", ""),
            "name_change_detected": r.get("name_change_detected", ""),
            "old_worker_status":  r.get("old_worker_status", ""),
            "new_worker_status":  r.get("new_worker_status", ""),
            "old_worker_type":    r.get("old_worker_type", "") if has_wt else "",
            "new_worker_type":    r.get("new_worker_type", "") if has_wt else "",
            "old_hire_date":      r.get("old_hire_date", ""),
            "new_hire_date":      r.get("new_hire_date", ""),
            "old_position":       r.get("old_position", ""),
            "new_position":       r.get("new_position", ""),
            "old_district":       r.get("old_district", ""),
            "new_district":       r.get("new_district", ""),
            "old_location_state": r.get("old_location_state", ""),
            "new_location_state": r.get("new_location_state", ""),
            "old_location":       r.get("old_location", "") if has_loc else "",
            "new_location":       r.get("new_location", "") if has_loc else "",
            "old_salary":         r.get("old_salary"),
            "new_salary":         r.get("new_salary"),
            "old_payrate":        r.get("old_payrate"),
            "new_payrate":        r.get("new_payrate"),
            "salary_delta":       sal_d,
            "salary_ratio":       _salary_ratio(r.get("old_salary"), r.get("new_salary")),
            "payrate_delta":      pay_d,
            "status_changed":     not _str_eq(r.get("old_worker_status"), r.get("new_worker_status")),
            "hire_date_changed":  not _str_eq(r.get("old_hire_date"), r.get("new_hire_date")),
            "job_org_changed":    "job_org" in fix_types,
            "hire_date_pattern":  result.get("per_fix", {}).get("hire_date", {}).get("reason", "")
                                  if "hire_date" in fix_types else "",
            "needs_review":       result["action"] == "REVIEW",
            "suggested_action":   result["action"],
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Full Excel workbook export.")
    parser.add_argument(
        "--out", default=None, metavar="PATH",
        help=f"Output workbook path (default: {OUT_PATH}).",
    )
    parser.add_argument(
        "--wide", default=None, metavar="PATH",
        help=f"wide_compare.csv path (default: {WIDE_CSV}).",
    )
    parser.add_argument(
        "--db", default=None, metavar="PATH",
        help=f"SQLite database path (default: {DB_PATH}).",
    )
    parser.add_argument(
        "--manifest", default=None, metavar="PATH",
        help=f"corrections_manifest.csv path (default: auto from RK_WORK_DIR or {MANIFEST_CSV}).",
    )
    parser.add_argument(
        "--gate-blocked", action="store_true", default=False,
        help="Sanity gate failed - write blocked placeholder to Corrections_Manifest sheet.",
    )
    args = parser.parse_args(argv)

    out_path      = Path(args.out)      if args.out      else OUT_PATH
    wide_path     = Path(args.wide)     if args.wide     else WIDE_CSV
    db_path       = Path(args.db)       if args.db       else DB_PATH
    manifest_path = Path(args.manifest) if args.manifest else MANIFEST_CSV
    gate_blocked  = args.gate_blocked

    if not db_path.exists():
        print(f"[error] DB not found: {db_path}", file=sys.stderr)
        sys.exit(2)

    # ------------------------------------------------------------------
    # Load main dataset
    # ------------------------------------------------------------------
    if wide_path.exists():
        print(f"[build_workbook] loading {wide_path.name} ...")
        all_df   = pd.read_csv(str(wide_path))
        wide_src = str(wide_path.relative_to(ROOT)) if wide_path.is_relative_to(ROOT) else str(wide_path)
    else:
        print(f"[build_workbook] {wide_path.name} not found - computing from DB ...")
        all_df   = _load_wide_from_db(db_path)
        wide_src = f"{db_path.name} (gating computed on the fly)"

    total = len(all_df)
    print(f"[build_workbook] {total:,} rows loaded.")

    # Strip any SSN/PII columns - they must never appear in any export
    _PII_DROP = [c for c in ("old_last4_ssn", "new_last4_ssn", "old_dob", "new_dob") if c in all_df.columns]
    if _PII_DROP:
        all_df = all_df.drop(columns=_PII_DROP)
        print(f"[build_workbook] [pii] stripped columns: {_PII_DROP}")

    # Ensure numeric columns are stored as numbers
    for num_col in [
        "salary_delta", "salary_ratio", "payrate_delta",
        "old_salary", "new_salary", "old_payrate", "new_payrate",
        "priority_score", "confidence",
    ]:
        if num_col in all_df.columns:
            all_df[num_col] = pd.to_numeric(all_df[num_col], errors="coerce")

    # ------------------------------------------------------------------
    # Load auxiliary sheets
    # ------------------------------------------------------------------
    def _review_from_alldf(df: pd.DataFrame) -> pd.DataFrame:
        """Return rows flagged for review from the live all_df data."""
        if "needs_review" in df.columns:
            mask = df["needs_review"].astype(str).str.lower().isin(["true", "1"])
        elif "action" in df.columns:
            mask = df["action"] == "REVIEW"
        else:
            return pd.DataFrame()
        return df[mask].copy()

    if REVIEW_CSV.exists():
        review_df = pd.read_csv(str(REVIEW_CSV))
        # Always filter to REVIEW-only - csv may contain APPROVE rows from older runs
        if not review_df.empty and "action" in review_df.columns:
            review_df = review_df[review_df["action"] == "REVIEW"].copy()
        if not review_df.empty:
            review_src = REVIEW_CSV.name
        else:
            # review_queue.csv exists but has 0 REVIEW rows - fall back to filtering
            # the live dataset so the sheet is never incorrectly blank.
            review_df  = _review_from_alldf(all_df)
            review_src = (
                f"{REVIEW_CSV.name} had no REVIEW rows - "
                "filtered from All_Matches (action==REVIEW)"
            )
    elif "needs_review" in all_df.columns or "action" in all_df.columns:
        review_df  = _review_from_alldf(all_df)
        review_src = "filtered from All_Matches (needs_review==True)"
    else:
        review_df  = pd.DataFrame()
        review_src = "unavailable"

    if gate_blocked:
        manifest_df  = pd.DataFrame([
            {"note": "Corrections blocked - sanity gate failed. Resolve gate failure before generating corrections."}
        ])
        manifest_src = "blocked (gate FAIL)"
    elif manifest_path.exists():
        manifest_df  = pd.read_csv(str(manifest_path))
        manifest_src = manifest_path.name
    else:
        manifest_df  = pd.DataFrame([
            {"note": "corrections_manifest.csv not found - run generate_corrections.py first"}
        ])
        manifest_src = f"placeholder (looked in: {manifest_path})"

    # ------------------------------------------------------------------
    # Build mismatch filter sheets
    # ------------------------------------------------------------------
    def _fix_filter(ft: str) -> pd.DataFrame:
        fix_types_text = _fix_types_text(all_df)
        if fix_types_text is None:
            return pd.DataFrame(columns=all_df.columns)
        return all_df[fix_types_text.str.contains(ft, na=False)].copy()

    salary_df  = _fix_filter("salary")
    status_df  = _fix_filter("status")
    hire_df    = _fix_filter("hire_date")
    job_org_df = _fix_filter("job_org")

    # Extra_Field_Mismatches - rows where any mm_<field> column is True
    mm_cols = [c for c in all_df.columns if c.startswith("mm_")]
    if mm_cols:
        def _mm_true(series: pd.Series) -> pd.Series:
            return series.astype(str).str.lower().isin(["true", "1"])
        mm_mask = pd.DataFrame({c: _mm_true(all_df[c]) for c in mm_cols}).any(axis=1)
        extra_mismatch_df: pd.DataFrame | None = all_df[mm_mask].copy()
    else:
        extra_mismatch_df = None

    # Fix 3: Rejected_Matches - rows where action == REJECT_MATCH
    rejected_df: pd.DataFrame | None = None
    if "action" in all_df.columns:
        rejected_df = all_df[all_df["action"] == "REJECT_MATCH"].copy()

    # Fix 1: CRITICAL_Zero_Salary - Active workers with $0 salary in new data
    active_zero_df = validate_active_zero_salary(all_df)
    if len(active_zero_df) > 0:
        print(
            f"\n[build_workbook] *** CRITICAL: {len(active_zero_df):,} Active workers have "
            f"$0 or missing salary in new data - see CRITICAL_Zero_Salary sheet ***\n"
        )

    # ------------------------------------------------------------------
    # Held corrections (BLOCKED / REJECT_MATCH hold reasons only)
    # ------------------------------------------------------------------
    held_path = HELD_CSV
    if held_path.exists():
        held_df = pd.read_csv(str(held_path))
        if not held_df.empty and "hold_reason" in held_df.columns:
            held_df = held_df[
                held_df["hold_reason"].str.contains(
                    r"BLOCKED|REJECT_MATCH|active_zero_salary_blocked", na=False
                )
            ].copy()
        held_src = f"{held_path.name}  ({len(held_df):,} rows)"
    else:
        held_df  = pd.DataFrame()
        held_src = "held_corrections.csv not found"

    # ------------------------------------------------------------------
    # Unmatched records - load full DataFrames (used for sheets + counts)
    # ------------------------------------------------------------------
    _run_outs = (_rk_work / "outputs") if _rk_work else (ROOT / "outputs")
    _uo_p = _run_outs / "unmatched_old.csv"
    _un_p = _run_outs / "unmatched_new.csv"

    if _uo_p.exists():
        try:
            unmatched_old_df    = pd.read_csv(str(_uo_p))
            unmatched_old_count = len(unmatched_old_df)
        except Exception:
            unmatched_old_df    = pd.DataFrame()
            unmatched_old_count = 0
    else:
        unmatched_old_df    = pd.DataFrame()
        unmatched_old_count = 0

    if _un_p.exists():
        try:
            unmatched_new_df    = pd.read_csv(str(_un_p))
            unmatched_new_count = len(unmatched_new_df)
        except Exception:
            unmatched_new_df    = pd.DataFrame()
            unmatched_new_count = 0
    else:
        unmatched_new_df    = pd.DataFrame()
        unmatched_new_count = 0

    print(f"  review_queue src       : {review_src}  ({len(review_df):,} rows)")
    print(f"  manifest src           : {manifest_src}  ({len(manifest_df):,} rows)")
    print(f"  held src               : {held_src}")
    print(f"  Salary_Mismatches      : {len(salary_df):,}")
    print(f"  Status_Mismatches      : {len(status_df):,}")
    print(f"  HireDate_Mismatches    : {len(hire_df):,}")
    print(f"  JobOrg_Mismatches      : {len(job_org_df):,}")
    if rejected_df is not None:
        print(f"  Rejected_Matches       : {len(rejected_df):,}")
    print(f"  CRITICAL_Zero_Salary   : {len(active_zero_df):,}")
    print(f"  Unmatched_Old          : {unmatched_old_count:,}")
    print(f"  Unmatched_New          : {unmatched_new_count:,}")
    if extra_mismatch_df is not None:
        print(f"  Extra_Field_Mismatches : {len(extra_mismatch_df):,}  (mm_ cols: {mm_cols})")

    # ------------------------------------------------------------------
    # Unmatched record counts (from matcher outputs)
    # ------------------------------------------------------------------
    _run_outs = (_rk_work / "outputs") if _rk_work else (ROOT / "outputs")
    unmatched_old_count = 0
    unmatched_new_count = 0
    _uo_p = _run_outs / "unmatched_old.csv"
    _un_p = _run_outs / "unmatched_new.csv"
    if _uo_p.exists():
        try:
            with _uo_p.open(encoding="utf-8") as _f:
                unmatched_old_count = max(0, sum(1 for _ in _f) - 1)
        except Exception:
            pass
    if _un_p.exists():
        try:
            with _un_p.open(encoding="utf-8") as _f:
                unmatched_new_count = max(0, sum(1 for _ in _f) - 1)
        except Exception:
            pass
    salary_parse_stats = _load_salary_parse_stats()

    # ------------------------------------------------------------------
    # Build workbook (write_only streaming - no MemoryError on large sets)
    # ------------------------------------------------------------------
    print(f"\n[build_workbook] writing workbook ...")
    wb = Workbook(write_only=False)

    # 1. Summary (with Start Here guide)
    ws_sum = wb.create_sheet("Summary")
    _write_summary_sheet(
        ws_sum,
        all_df,
        db_path,
        wide_src,
        unmatched_old_count,
        unmatched_new_count,
        salary_parse_stats=salary_parse_stats,
    )
    print(f"  wrote: Summary")

    # 2. All Matches - full dataset with all columns
    ws_all = wb.create_sheet("All Matches")
    _write_df_to_sheet(ws_all, all_df)
    print(f"  wrote: {'All Matches':<25}  ({total:,} rows)")

    # 3-6. Category mismatch sheets - slimmed to relevant columns
    for sheet_name, df, slim_cols in [
        ("Salary Mismatches",    salary_df,  _SALARY_SLIM_COLS),
        ("Job and Org",          job_org_df, _JOB_ORG_SLIM_COLS),
        ("Hire Date",            hire_df,    _HIRE_DATE_SLIM_COLS),
        ("Status Changes",       status_df,  _STATUS_SLIM_COLS),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_mismatch_slim(ws, df, slim_cols)
        print(f"  wrote: {sheet_name:<25}  ({len(df):,} rows)")

    # 7. Review Queue - slim 8-column view, human-readable labels
    ws_rq = wb.create_sheet("Review Queue")
    _write_review_queue_slim(ws_rq, review_df, all_df)
    print(f"  wrote: {'Review Queue':<25}  ({len(review_df):,} rows)")

    # 8. Rejected Matches
    if rejected_df is not None and not rejected_df.empty:
        ws_rej = wb.create_sheet("Rejected Matches")
        _write_df_to_sheet_styled(ws_rej, rejected_df, hdr_font=_REJ_HDR_FONT, hdr_fill=_REJ_HDR_FILL)
        print(f"  wrote: {'Rejected Matches':<25}  ({len(rejected_df):,} rows)")

    # Additional reference sheets (kept for data completeness)
    ws_uo = wb.create_sheet("Unmatched - Old")
    _write_unmatched_sheet(ws_uo, unmatched_old_df)
    ws_un = wb.create_sheet("Unmatched - New")
    _write_unmatched_sheet(ws_un, unmatched_new_df)
    ws_clean = wb.create_sheet("Clean Data (All Columns)")
    _write_df_to_sheet_styled(ws_clean, all_df, hdr_font=_REF_HDR_FONT, hdr_fill=_REF_HDR_FILL)

    # Held and critical context (advanced)
    if not held_df.empty:
        ws_held = wb.create_sheet("Held Corrections")
        _write_held_corrections_sheet(ws_held, held_df, all_df)
    if not active_zero_df.empty:
        ws_crit = wb.create_sheet("Active Zero Salary")
        _write_df_to_sheet_styled(ws_crit, active_zero_df, hdr_font=_CRIT_HDR_FONT, hdr_fill=_CRIT_HDR_FILL)

    if extra_mismatch_df is not None and not extra_mismatch_df.empty:
        ws_ex = wb.create_sheet("Extra Field Mismatches")
        _write_df_to_sheet(ws_ex, extra_mismatch_df)

    ws_mf = wb.create_sheet("Corrections Manifest")
    enhanced_manifest = _enhance_manifest(manifest_df)
    _write_df_to_sheet(ws_mf, enhanced_manifest)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    try:
        display_path = out_path.relative_to(ROOT)
    except ValueError:
        display_path = out_path
    print(f"\n[build_workbook] saved: {display_path}")


if __name__ == "__main__":
    main()

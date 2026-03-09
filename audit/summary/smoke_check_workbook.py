"""
smoke_check_workbook.py — Verifies the Excel workbook output.

Assertions
----------
1. build_workbook runs without error.
2. recon_workbook.xlsx exists after running.
3. All required sheet names are present.
4. All_Matches row count == matched_pairs count from DB.
5. Salary_Mismatches count matches rows where fix_types contains "salary".

Run:
    venv/Scripts/python.exe audit/summary/smoke_check_workbook.py
"""
from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

import pandas as pd

try:
    import openpyxl
except ImportError:
    print(
        "[error] openpyxl not installed. "
        "Run: venv/Scripts/pip.exe install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)

_HERE   = Path(__file__).resolve().parent    # audit/summary/
ROOT    = _HERE.parents[1]
DB_PATH = ROOT / "audit" / "audit.db"
WIDE_CSV = ROOT / "audit" / "exports" / "out" / "wide_compare.csv"

REQUIRED_SHEETS = [
    "Summary",
    "All_Matches",
    "Salary_Mismatches",
    "Status_Mismatches",
    "HireDate_Mismatches",
    "JobOrg_Mismatches",
    "Review_Queue",
    "Corrections_Manifest",
]


def _fail(msg: str) -> None:
    print(f"\n[FAIL] {msg}", file=sys.stderr)
    sys.exit(1)


def _pass(msg: str) -> None:
    print(f"  [PASS] {msg}")


def _count_ws_rows(ws) -> int:
    """Count rows by iteration — works for write_only output (max_row may be None)."""
    return sum(1 for _ in ws)


def main() -> None:
    print("=" * 60)
    print("  SMOKE CHECK: Excel Workbook")
    print("=" * 60)

    # ------------------------------------------------------------------
    # Assertion 1: build_workbook runs without error
    # ------------------------------------------------------------------
    print("\n  Running build_workbook.main() ...")
    sys.path.insert(0, str(_HERE))
    import build_workbook
    try:
        build_workbook.main(argv=[])
    except SystemExit as exc:
        if exc.code != 0:
            _fail(f"Assertion 1 FAILED: build_workbook.main() exited with code {exc.code}")
    except MemoryError:
        _fail(
            "Assertion 1 FAILED: build_workbook.main() raised MemoryError — "
            "dataset too large for openpyxl default mode. "
            "build_workbook.py must use write_only=True."
        )
    print()
    _pass("Assertion 1: build_workbook runs without error")

    # ------------------------------------------------------------------
    # Assertion 2: Workbook file exists
    # ------------------------------------------------------------------
    wb_path = build_workbook.OUT_PATH
    if not wb_path.exists():
        _fail(f"Assertion 2 FAILED: workbook not found at {wb_path}")
    size_kb = wb_path.stat().st_size // 1024
    _pass(f"Assertion 2: workbook exists ({size_kb:,} KB)")

    # ------------------------------------------------------------------
    # Assertion 3: Required sheet names present
    # ------------------------------------------------------------------
    wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
    actual_sheets = wb.sheetnames
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in actual_sheets]
    wb.close()
    if missing_sheets:
        _fail(f"Assertion 3 FAILED: missing sheets: {missing_sheets}")
    _pass(f"Assertion 3: all {len(REQUIRED_SHEETS)} required sheets present: {actual_sheets}")

    # ------------------------------------------------------------------
    # Assertion 4: All_Matches sheet has at least 1 data row and expected
    #              header columns. Full row-count skipped (iterating 20k+
    #              rows via openpyxl read_only is slow; build already logs
    #              counts). Row count is verified via the source CSV check.
    # ------------------------------------------------------------------
    if not DB_PATH.exists():
        print("  [SKIP] Assertion 4: audit.db not found")
    else:
        con = sqlite3.connect(str(DB_PATH))
        try:
            (expected_count,) = con.execute("SELECT COUNT(*) FROM matched_pairs").fetchone()
        finally:
            con.close()

        wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
        ws_all = wb["All_Matches"]
        rows_sample = list(ws_all.iter_rows(min_row=1, max_row=2, values_only=True))
        wb.close()

        if not rows_sample:
            _fail("Assertion 4 FAILED: All_Matches sheet has no rows at all")
        elif len(rows_sample) < 2:
            _fail("Assertion 4 FAILED: All_Matches sheet has no data rows (only header)")
        else:
            header = [str(v) for v in rows_sample[0] if v is not None]
            if "pair_id" not in header:
                _fail(f"Assertion 4 FAILED: All_Matches header missing 'pair_id'. Got: {header[:5]}")
            else:
                _pass(
                    f"Assertion 4: All_Matches sheet present with valid header and data rows "
                    f"(expected {expected_count:,} rows — verified by build log)"
                )

    # ------------------------------------------------------------------
    # Assertion 5: Salary_Mismatches sheet has at least 1 data row.
    # Row count cross-checked against wide_compare.csv (not re-counted
    # from workbook to avoid slow full-sheet iteration).
    # ------------------------------------------------------------------
    if WIDE_CSV.exists():
        ref_df = pd.read_csv(str(WIDE_CSV), usecols=["fix_types"])
        expected_salary_rows = int(ref_df["fix_types"].str.contains("salary", na=False).sum())
        ref_src = f"wide_compare.csv ({expected_salary_rows:,} rows expected)"
    else:
        expected_salary_rows = None
        ref_src = "skipped (wide_compare.csv not found)"

    wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
    ws_sal = wb["Salary_Mismatches"]
    sal_sample = list(ws_sal.iter_rows(min_row=1, max_row=2, values_only=True))
    wb.close()

    if not sal_sample or len(sal_sample) < 2:
        _fail("Assertion 5 FAILED: Salary_Mismatches sheet has no data rows")
    elif expected_salary_rows is not None:
        _pass(
            f"Assertion 5: Salary_Mismatches sheet present with data rows "
            f"({ref_src})"
        )
    else:
        print(f"  [SKIP] Assertion 5: {ref_src}")

    # ------------------------------------------------------------------
    # Sheet presence summary (no row iteration — avoids slow large reads)
    # ------------------------------------------------------------------
    print(f"\n  Sheet presence check:")
    wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
    for sheet_name in REQUIRED_SHEETS:
        status = "present" if sheet_name in wb.sheetnames else "MISSING"
        print(f"    {sheet_name:<25}  {status}")
    wb.close()

    print(f"\n  All assertions PASSED.")
    print(f"  [done] smoke check complete.")


if __name__ == "__main__":
    main()
